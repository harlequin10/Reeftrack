import os
import json
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.db import transaction
from django.utils import timezone
from django.http import JsonResponse
from .forms import RegisterForm, LoginForm, AdminCreateUserForm, UserProfileForm, CustomPasswordChangeForm
from .decorators import admin_required, curator_required, contributor_required
from .models import (
    UserProfile, Municipality, Barangay, Assessment,
    Transect, Species, TransectSpecies, AssessmentImage, Contributor, CustomMethodology,
    BarangayTransect
)

def home(request):
    """Public home page"""
    recent_assessments = Assessment.objects.filter(
        status='approved'
    ).select_related('municipality', 'barangay').order_by('-approved_at')[:6]

    return render(request, 'public/index.html', {
        'recent_assessments': recent_assessments,
        'active_page': 'home',
    })

def public_assessment_detail(request, assessment_id):
    """Public: View approved assessment details (no login required)."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects__species_data__species', 'images'),
        id=assessment_id,
        status='approved',
    )
    transects = assessment.transects.prefetch_related('species_data__species').all()

    parsed_species = []
    has_species_records = TransectSpecies.objects.filter(transect__assessment=assessment).exists()
    if not has_species_records:
        for transect in transects:
            if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
                sp, _ = parse_cpc_excel(transect.shallow_excel.path)
                for s in sp:
                    s['depth'] = 'Shallow'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)
            if transect.deep_excel and os.path.exists(transect.deep_excel.path):
                sp, _ = parse_cpc_excel(transect.deep_excel.path)
                for s in sp:
                    s['depth'] = 'Deep'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)

    return render(request, 'public/assessment_detail.html', {
        'assessment': assessment,
        'transects': transects,
        'parsed_species': parsed_species,
        'has_species_records': has_species_records,
    })

def about(request):
    """Public about page"""
    return render(request, 'public/about.html', {'active_page': 'about'})

def register(request):
    """Registration page - Only allows contributor role"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(
                request, 
                f'Registration successful! Welcome {user.username}! '
                'Your account is pending approval. '
                'You will be notified once an admin approves your account.'
            )
            return redirect('login')
        else:
            error_list = []
            for field, errors in form.errors.items():
                label = form[field].label if field in form.fields else field.replace('_', ' ').title()
                for err in errors:
                    error_list.append(f"{label}: {err}")
            messages.error(request, '\n'.join(error_list) if error_list else 'Please correct the errors below.')
    else:
        form = RegisterForm()
    
    return render(request, 'public/register.html', {'form': form})

def login_view(request):
    """Public login page - email + password"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password')
            try:
                user_obj = User.objects.get(email=email)
                user = authenticate(username=user_obj.username, password=password)
            except User.DoesNotExist:
                user = None
            
            if user is not None:
                if not user.is_active:
                    messages.error(request, 'Your account has been deactivated. Please contact an administrator.')
                    return render(request, 'public/login.html', {'form': form})
                
                if hasattr(user, 'profile'):
                    if user.profile.status == 'approved':
                        login(request, user)
                        messages.success(request, f'Welcome back {user.profile.get_full_name() or user.email}!')
                        return redirect('dashboard')
                    elif user.profile.status == 'pending':
                        messages.warning(
                            request, 
                            'Your account is pending approval. '
                            'Please wait for an admin or curator to approve your account.'
                        )
                    elif user.profile.status == 'rejected':
                        messages.error(
                            request, 
                            'Your account has been rejected. '
                            f'Reason: {user.profile.rejection_reason or "No reason provided"}'
                        )
                else:
                    messages.error(request, 'Invalid account configuration.')
            else:
                messages.error(request, 'Invalid email or password.')
        else:
            messages.error(request, 'Invalid email or password.')
    else:
        form = LoginForm()
    
    return render(request, 'public/login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('home')

@login_required
def dashboard(request):
    """Redirect to role-specific dashboard"""
    user_role = request.user.profile.role if hasattr(request.user, 'profile') else 'contributor'
    
    if user_role == 'admin':
        return redirect('admin_dashboard')
    elif user_role == 'curator':
        return redirect('curator_dashboard')
    else:
        return redirect('contributor_dashboard')

@login_required
@admin_required
def admin_dashboard(request):
    """Admin dashboard with user stats and assessment stats"""
    total_users = User.objects.count()
    
    # Only count pending contributors (admins and curators are auto-approved)
    pending_count = UserProfile.objects.filter(role='contributor', status='pending').count()
    approved_count = UserProfile.objects.filter(status='approved').count()
    
    # Assessment stats
    pending_assessments = Assessment.objects.filter(status='submitted').count()
    total_assessments = Assessment.objects.count()
    approved_assessments = Assessment.objects.filter(status='approved').count()
    
    # Location & species stats
    total_municipalities = Municipality.objects.count()
    total_barangays = Barangay.objects.count()
    total_species = Species.objects.count()
    
    # Recent approved assessments
    recent_approved = Assessment.objects.filter(
        status='approved'
    ).select_related('municipality', 'barangay', 'uploaded_by').order_by('-approved_at')[:5]
    
    context = {
        'total_users': total_users,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'pending_assessments': pending_assessments,
        'total_assessments': total_assessments,
        'approved_assessments': approved_assessments,
        'total_municipalities': total_municipalities,
        'total_barangays': total_barangays,
        'total_species': total_species,
        'recent_approved': recent_approved,
    }
    
    return render(request, 'admin/dashboard.html', context)

@login_required
@admin_required
def admin_manage_users(request):
    """
    Admin view to manage all users with dropdown actions
    Admins cannot see edit/delete options for other admins
    """
    users = User.objects.all().select_related('profile').order_by('-date_joined')
    
    # Get filter parameters
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Apply filters
    if role_filter:
        users = users.filter(profile__role=role_filter)
    if status_filter:
        users = users.filter(profile__status=status_filter)
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) | 
            Q(email__icontains=search_query)
        )
    
    context = {
        'users': users,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'search_query': search_query,
        'role_choices': UserProfile.ROLE_CHOICES,
        'status_choices': UserProfile.STATUS_CHOICES,
        'is_superuser': request.user.is_superuser,
    }
    
    return render(request, 'admin/manage_users/index.html', context)

@login_required
@admin_required
def admin_user_action(request, user_id):
    """
    Handle all user actions via dropdown for Admin
    Cannot perform actions on other admins
    """
    user = get_object_or_404(User, id=user_id)
    
    if user == request.user:
        messages.error(request, 'You cannot perform actions on yourself!')
        return redirect('admin_manage_users')
    
    # Prevent actions on other admins (except superuser can manage all)
    if user.profile.role == 'admin' and not request.user.is_superuser:
        messages.error(request, 'You cannot perform actions on other admin accounts!')
        return redirect('admin_manage_users')
    
    action = request.POST.get('action')
    
    if action == 'approve':
        if user.profile.role == 'contributor' and user.profile.status == 'pending':
            user.profile.status = 'approved'
            user.profile.approved_by = request.user
            user.profile.approved_at = timezone.now()
            user.is_active = True
            user.save()
            user.profile.save()
            messages.success(request, f'✅ {user.username} has been approved!')
        else:
            messages.warning(request, f'{user.username} is not a pending contributor.')
    
    elif action == 'reject':
        if user.profile.role == 'contributor' and user.profile.status == 'pending':
            rejection_reason = request.POST.get('rejection_reason', 'No reason provided')
            user.profile.status = 'rejected'
            user.profile.rejection_reason = rejection_reason
            user.is_active = False
            user.save()
            user.profile.save()
            messages.info(request, f'❌ {user.username} has been rejected.')
        else:
            messages.warning(request, f'{user.username} is not a pending contributor.')
    
    elif action == 'activate':
        user.is_active = True
        user.save()
        messages.success(request, f'✅ {user.username} has been activated!')
    
    elif action == 'deactivate':
        user.is_active = False
        user.save()
        messages.warning(request, f'⛔ {user.username} has been deactivated.')
    
    elif action == 'delete':
        if user.profile.role == 'admin' and not request.user.is_superuser:
            messages.error(request, 'You cannot delete other admin accounts!')
            return redirect('admin_manage_users')
        username = user.username
        user.delete()
        messages.success(request, f'✅ {username} has been deleted.')
    
    else:
        messages.error(request, 'Invalid action selected.')
    
    return redirect('admin_manage_users')

@login_required
@admin_required
def admin_edit_user(request, user_id):
    """
    Admin view to edit user details - Cannot edit other admins
    """
    user = get_object_or_404(User, id=user_id)
    
    # Prevent editing own account
    if user == request.user:
        messages.error(request, 'You cannot edit your own account!')
        return redirect('admin_manage_users')
    
    # Prevent editing other admins (except superuser)
    if user.profile.role == 'admin' and not request.user.is_superuser:
        messages.error(request, 'You cannot edit other admin accounts!')
        return redirect('admin_manage_users')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        middle_initial = request.POST.get('middle_initial', '').strip()[:1].upper()
        suffix = request.POST.get('suffix', '')
        email = request.POST.get('email')
        role = request.POST.get('role')
        status = request.POST.get('status')
        
        user.first_name = first_name
        user.last_name = last_name
        if email:
            user.email = email
        
        user.profile.middle_initial = middle_initial
        user.profile.suffix = suffix
        
        if role and role in dict(UserProfile.ROLE_CHOICES):
            if role == 'admin' and not request.user.is_superuser:
                messages.error(request, 'Only superusers can assign admin role!')
                return redirect('admin_manage_users')
            user.profile.role = role
            if role in ['admin', 'curator']:
                user.profile.status = 'approved'
                user.is_active = True
        
        if status and status in dict(UserProfile.STATUS_CHOICES):
            if user.profile.role == 'contributor':
                user.profile.status = status
                if status == 'approved':
                    user.profile.approved_by = request.user
                    user.profile.approved_at = timezone.now()
                    user.is_active = True
                elif status == 'rejected':
                    user.is_active = False
        
        user.save()
        user.profile.save()
        
        # Superuser-only password reset
        if request.user.is_superuser:
            new_password = request.POST.get('new_password', '').strip()
            confirm_password = request.POST.get('confirm_password', '').strip()
            if new_password:
                if len(new_password) < 8:
                    messages.error(request, 'Password must be at least 8 characters.')
                    return redirect('admin_edit_user', user_id=user.id)
                if new_password != confirm_password:
                    messages.error(request, 'Passwords do not match.')
                    return redirect('admin_edit_user', user_id=user.id)
                user.set_password(new_password)
                user.save()
                messages.success(request, f'Password updated for {user.email}.')
        
        messages.success(request, f'User has been updated successfully!')
        return redirect('admin_manage_users')
    
    context = {
        'user': user,
        'role_choices': UserProfile.ROLE_CHOICES,
        'status_choices': UserProfile.STATUS_CHOICES,
        'is_superuser': request.user.is_superuser,
    }
    
    return render(request, 'admin/manage_users/edit.html', context)

@login_required
@admin_required
def admin_create_user(request):
    """
    Admin view to create users with any role
    Only superusers can create admin accounts
    """
    if request.method == 'POST':
        form = AdminCreateUserForm(request.POST)
        if form.is_valid():
            # Check if trying to create an admin without superuser privileges
            if form.cleaned_data['role'] == 'admin' and not request.user.is_superuser:
                messages.error(request, 'Only superusers can create admin accounts!')
                return redirect('admin_manage_users')
            
            user = form.save()
            role_display = dict(UserProfile.ROLE_CHOICES).get(form.cleaned_data['role'])
            messages.success(request, f'User created successfully with role: {role_display} (Auto-approved)')
            return redirect('admin_manage_users')
        else:
            error_list = []
            for field, errors in form.errors.items():
                label = form[field].label if field in form.fields else field.replace('_', ' ').title()
                for err in errors:
                    error_list.append(f"{label}: {err}")
            messages.error(request, '\n'.join(error_list) if error_list else 'Please correct the errors below.')
    else:
        form = AdminCreateUserForm()
    
    context = {
        'form': form,
        'is_superuser': request.user.is_superuser,
    }
    
    return render(request, 'admin/manage_users/create.html', context)

# ==================== CURATOR VIEWS ====================

@login_required
@curator_required
def curator_dashboard(request):
    """Curator dashboard - Overview stats and quick actions"""
    contributors = User.objects.filter(profile__role='contributor')
    total_contributors = contributors.count()
    pending_count = contributors.filter(profile__status='pending').count()
    approved_count = contributors.filter(profile__status='approved').count()
    rejected_count = contributors.filter(profile__status='rejected').count()

    # Assessment stats
    pending_assessments = Assessment.objects.filter(status='submitted').count()
    total_assessments = Assessment.objects.count()

    context = {
        'total_contributors': total_contributors,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'pending_assessments': pending_assessments,
        'total_assessments': total_assessments,
    }
    return render(request, 'curator/dashboard.html', context)

@login_required
@curator_required
def curator_manage_contributors(request):
    """Manage contributors - View, filter, approve, reject"""
    contributors = User.objects.filter(profile__role='contributor').select_related('profile').order_by('-date_joined')
    total_contributors = contributors.count()
    pending_count = contributors.filter(profile__status='pending').count()
    approved_count = contributors.filter(profile__status='approved').count()
    rejected_count = contributors.filter(profile__status='rejected').count()

    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    if status_filter:
        contributors = contributors.filter(profile__status=status_filter)
    if search_query:
        contributors = contributors.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    context = {
        'contributors': contributors,
        'total_contributors': total_contributors,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'status_filter': status_filter,
        'search_query': search_query,
        'status_choices': UserProfile.STATUS_CHOICES,
    }
    return render(request, 'curator/manage_contributors.html', context)

@login_required
@curator_required
def curator_contributor_action(request, user_id):
    """
    Curator view to handle contributor actions (approve, reject, activate, deactivate)
    """
    user = get_object_or_404(User, id=user_id)
    
    # Only allow actions on contributors
    if user.profile.role != 'contributor':
        messages.warning(request, 'You can only manage contributors!')
        return redirect('curator_dashboard')
    
    if user == request.user:
        messages.error(request, 'You cannot perform actions on yourself!')
        return redirect('curator_dashboard')
    
    action = request.POST.get('action')
    
    if action == 'approve':
        if user.profile.status == 'pending':
            user.profile.status = 'approved'
            user.profile.approved_by = request.user
            user.profile.approved_at = timezone.now()
            user.is_active = True
            user.save()
            user.profile.save()
            messages.success(request, f'✅ {user.username} has been approved!')
        else:
            messages.warning(request, f'{user.username} is not pending approval.')
    
    elif action == 'reject':
        if user.profile.status == 'pending':
            rejection_reason = request.POST.get('rejection_reason', 'No reason provided')
            user.profile.status = 'rejected'
            user.profile.rejection_reason = rejection_reason
            user.is_active = False
            user.save()
            user.profile.save()
            messages.info(request, f'❌ {user.username} has been rejected.')
        else:
            messages.warning(request, f'{user.username} is not pending approval.')
    
    elif action == 'activate':
        user.is_active = True
        user.save()
        messages.success(request, f'✅ {user.username} has been activated!')
    
    elif action == 'deactivate':
        user.is_active = False
        user.save()
        messages.warning(request, f'⛔ {user.username} has been deactivated.')
    
    else:
        messages.error(request, 'Invalid action selected.')
    
    return redirect('curator_dashboard')

@login_required
@curator_required
def curator_edit_contributor(request, user_id):
    """
    Curator view to edit contributor details
    """
    user = get_object_or_404(User, id=user_id)
    
    # Only allow editing contributors
    if user.profile.role != 'contributor':
        messages.warning(request, 'You can only edit contributors!')
        return redirect('curator_dashboard')
    
    if user == request.user:
        messages.error(request, 'You cannot edit your own account!')
        return redirect('curator_dashboard')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        middle_initial = request.POST.get('middle_initial', '').strip()[:1].upper()
        suffix = request.POST.get('suffix', '')
        email = request.POST.get('email')
        status = request.POST.get('status')
        
        user.first_name = first_name
        user.last_name = last_name
        if email:
            user.email = email
        user.profile.middle_initial = middle_initial
        user.profile.suffix = suffix
        
        if status and status in dict(UserProfile.STATUS_CHOICES):
            user.profile.status = status
            if status == 'approved':
                user.profile.approved_by = request.user
                user.profile.approved_at = timezone.now()
                user.is_active = True
            elif status == 'rejected':
                user.is_active = False
        
        user.save()
        user.profile.save()
        
        messages.success(request, f'Contributor has been updated successfully!')
        return redirect('curator_dashboard')
    
    context = {
        'user': user,
        'status_choices': UserProfile.STATUS_CHOICES,
    }
    
    return render(request, 'curator/edit_contributor.html', context)

@login_required
@contributor_required
def contributor_dashboard(request):
    """Contributor dashboard"""
    my_assessments = Assessment.objects.filter(uploaded_by=request.user).order_by('-created_at')[:5]
    total = Assessment.objects.filter(uploaded_by=request.user).count()
    return render(request, 'contributor/dashboard.html', {
        'my_assessments': my_assessments,
        'total_assessments': total,
    })


# ==================== ASSESSMENT UPLOAD VIEWS ====================

def parse_cpc_excel(filepath):
    """
    Parse a CPCe Excel file (one depth).
    Expected format: Sub Category | Major Category | Mean
    Returns (species_list, errors) tuple.
    species_list = [{'sub_category': ..., 'major_category': ..., 'cover': ...}, ...]
    errors = ['error message 1', ...]
    Only returns species with valid (numeric, > 0) cover values.
    """
    import openpyxl
    errors = []
    species_list = []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return [], [f'Could not open Excel file: {e}']

    # Find a usable sheet: try preferred names first, then fall back to first sheet
    target_sheet = None
    for name in ['whole'] + [n for n in wb.sheetnames if n.lower().startswith('transect')] + ['Sheet1', 'Sheet']:
        if name in wb.sheetnames:
            target_sheet = name
            break
    if not target_sheet and wb.sheetnames:
        target_sheet = wb.sheetnames[0]

    if not target_sheet:
        wb.close()
        return [], ['No usable sheet found in the Excel file.']

    ws = wb[target_sheet]

    # Read header row to validate columns
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip().lower() if h else '' for h in row]
        break

    # Validate required columns exist
    required = ['sub category', 'major category']
    missing = [r for r in required if r not in headers]
    if missing:
        wb.close()
        return [], [f'Missing required columns: {", ".join(missing)} in sheet "{target_sheet}". Your file must have columns: Sub Category, Major Category, Mean']

    # Find column indices
    try:
        sub_idx = headers.index('sub category')
        major_idx = headers.index('major category')
    except ValueError:
        wb.close()
        return [], ['Could not find Sub Category and Major Category columns.']

    mean_idx = None
    for candidate in ['mean', 'final mean', 'average', '%average', 'deep', 'shallow']:
        for i, h in enumerate(headers):
            if h == candidate:
                mean_idx = i
                break
        if mean_idx is not None:
            break

    if mean_idx is None:
        if len(headers) >= 3:
            mean_idx = 2
        else:
            wb.close()
            return [], ['No Mean/cover column found. File must have a Mean column (or deep/shallow).']

    # Read data rows
    skipped = 0
    bad_rows = []
    seen = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= max(sub_idx, major_idx, mean_idx):
            continue
        sub_val = row[sub_idx]
        major_val = row[major_idx]
        cover_val = row[mean_idx]

        if not sub_val or not major_val:
            continue
        sub_cat = str(sub_val).strip()
        major_cat = str(major_val).strip()
        if sub_cat.lower() in ('sub category', '') or major_cat.lower() in ('major category', ''):
            continue

        # Validate cover
        cover = None
        try:
            cover = float(cover_val)
        except (TypeError, ValueError):
            bad_rows.append(f'Row {row_num}: "{sub_cat}" has non-numeric cover value')
            skipped += 1
            continue

        if cover <= 0:
            bad_rows.append(f'Row {row_num}: "{sub_cat}" has cover {cover} (must be > 0)')
            skipped += 1
            continue

        # Deduplicate
        key = (sub_cat, major_cat)
        if key in seen:
            continue
        seen.add(key)

        species_list.append({
            'sub_category': sub_cat,
            'major_category': major_cat,
            'cover': round(cover, 2),
        })

    wb.close()

    # Build errors
    if bad_rows:
        errors.append(f'{len(bad_rows)} row(s) skipped (invalid data): ' + '; '.join(bad_rows[:5]))
        if len(bad_rows) > 5:
            errors.append(f'... and {len(bad_rows) - 5} more skipped rows')

    if not species_list:
        if not errors:
            errors.append('No valid species data found. File must have rows with Sub Category, Major Category, and a positive numeric Mean value.')
        return [], errors

    return species_list, errors


def save_barangay_transect_coords(assessment):
    """Save transect coordinates from an approved assessment to BarangayTransect records."""
    transects = assessment.transects.all()
    for t in transects:
        # Check if identical coordinates already exist for this barangay
        existing = BarangayTransect.objects.filter(
            barangay=assessment.barangay,
            shallow_start_lat=t.shallow_start_lat,
            shallow_start_lng=t.shallow_start_lng,
            shallow_end_lat=t.shallow_end_lat,
            shallow_end_lng=t.shallow_end_lng,
            deep_start_lat=t.deep_start_lat,
            deep_start_lng=t.deep_start_lng,
            deep_end_lat=t.deep_end_lat,
            deep_end_lng=t.deep_end_lng,
        ).first()
        if existing:
            # Update source/assessment reference if needed
            if existing.source != 'manual':
                existing.assessment = assessment
                existing.save()
        else:
            BarangayTransect.objects.create(
                barangay=assessment.barangay,
                shallow_start_lat=t.shallow_start_lat,
                shallow_start_lng=t.shallow_start_lng,
                shallow_end_lat=t.shallow_end_lat,
                shallow_end_lng=t.shallow_end_lng,
                deep_start_lat=t.deep_start_lat,
                deep_start_lng=t.deep_start_lng,
                deep_end_lat=t.deep_end_lat,
                deep_end_lng=t.deep_end_lng,
                source='assessment',
                assessment=assessment,
            )


def cleanup_barangay_transect_coords(assessment):
    """Remove BarangayTransect records from this assessment, but keep if another approved assessment uses them."""
    records = BarangayTransect.objects.filter(assessment=assessment, source='assessment')
    for rec in records:
        # Check if another approved assessment references this same barangay with identical coordinates
        other_uses = BarangayTransect.objects.filter(
            barangay=rec.barangay,
            source='assessment',
            shallow_start_lat=rec.shallow_start_lat,
            shallow_start_lng=rec.shallow_start_lng,
            shallow_end_lat=rec.shallow_end_lat,
            shallow_end_lng=rec.shallow_end_lng,
            deep_start_lat=rec.deep_start_lat,
            deep_start_lng=rec.deep_start_lng,
            deep_end_lat=rec.deep_end_lat,
            deep_end_lng=rec.deep_end_lng,
        ).exclude(assessment=assessment).filter(assessment__isnull=False, assessment__status='approved').exists()
        if other_uses:
            # Another approved assessment uses these coords — just detach
            rec.assessment = None
            rec.save(update_fields=['assessment'])
        else:
            rec.delete()


def populate_species_from_excel(assessment):
    """
    Parse Excel files for all transects in an assessment and create Species + TransectSpecies records.
    Called when an assessment is approved.
    """
    transects = assessment.transects.all()
    created_count = 0

    for transect in transects:
        # Parse shallow
        if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
            species_list, _ = parse_cpc_excel(transect.shallow_excel.path)
            for sp in species_list:
                cover = sp.get('cover', 0)
                if not cover or cover <= 0:
                    continue
                species, created = Species.objects.get_or_create(
                    sub_category=sp['sub_category'],
                    major_category=sp['major_category'],
                    defaults={'source': Species.SOURCE_ASSESSMENT},
                )
                TransectSpecies.objects.create(
                    transect=transect, species=species, depth='shallow', cover=cover
                )
                created_count += 1

        # Parse deep
        if transect.deep_excel and os.path.exists(transect.deep_excel.path):
            species_list, _ = parse_cpc_excel(transect.deep_excel.path)
            for sp in species_list:
                cover = sp.get('cover', 0)
                if not cover or cover <= 0:
                    continue
                species, created = Species.objects.get_or_create(
                    sub_category=sp['sub_category'],
                    major_category=sp['major_category'],
                    defaults={'source': Species.SOURCE_ASSESSMENT},
                )
                TransectSpecies.objects.create(
                    transect=transect, species=species, depth='deep', cover=cover
                )
                created_count += 1

    # Compute overall coral cover
    all_covers = TransectSpecies.objects.filter(
        transect__assessment=assessment
    ).values_list('cover', flat=True)
    if all_covers:
        assessment.overall_coral_cover = sum(Decimal(str(c)) for c in all_covers) / len(all_covers)
        assessment.compute_condition()
        assessment.save()

    return created_count


def check_duplicate_species(species_list, depth, pending_transects, barangay_id=None, assessment_id=None):
    """
    Check for duplicate species data.
    1. Within same assessment (other transects in session).
    2. Against previous assessments in DB (any barangay for exact, same barangay for partial).
    Returns list of warning strings.
    """
    if not species_list:
        return []

    warnings = []
    species_set = {(sp['sub_category'], sp['major_category'], round(sp['cover'], 2)) for sp in species_list}

    # 1. Check against other transects in the same assessment session
    for t in pending_transects:
        existing_key = f'shallow' if depth == 'shallow' else 'deep'
        other_key = f'deep' if depth == 'shallow' else 'shallow'
        existing_species = t.get(f'{existing_key}_species', [])
        if not existing_species:
            continue
        other_set = {(sp['sub_category'], sp['major_category'], round(sp['cover'], 2)) for sp in existing_species}
        overlap = species_set & other_set
        if len(overlap) == len(species_set) and len(species_set) == len(other_set) and len(species_set) > 0:
            warnings.append(
                f'Duplicate detected: This {depth} file has the exact same {len(species_set)} species with identical cover values as Transect {t["transect_number"]} ({existing_key}).'
            )
        elif len(overlap) > 0:
            warnings.append(
                f'{len(overlap)} of {len(species_set)} species match exactly with Transect {t["transect_number"]} ({existing_key}) (same species + cover).'
            )

    # 2. Check against previous assessments in DB
    # Check same barangay first, then other barangays
    if barangay_id:
        same_barangay = Assessment.objects.filter(
            barangay_id=barangay_id,
            status='approved',
        ).exclude(id=assessment_id or 0).order_by('-created_at')[:10]
    else:
        same_barangay = Assessment.objects.none()

    other_barangay = Assessment.objects.filter(
        status='approved',
    ).exclude(barangay_id=barangay_id).order_by('-created_at')[:10]

    checked_ids = set()

    for assessment in list(same_barangay) + list(other_barangay):
        if assessment.id in checked_ids:
            continue
        checked_ids.add(assessment.id)

        db_species = TransectSpecies.objects.filter(
            transect__assessment=assessment,
            depth=depth,
        ).select_related('species').values_list(
            'species__sub_category', 'species__major_category', 'cover'
        )
        db_set = {(s[0], s[1], round(float(s[2]), 2)) for s in db_species}
        if not db_set:
            continue

        overlap = species_set & db_set
        total = max(len(species_set), len(db_set))
        if total == 0:
            continue

        match_pct = (len(overlap) / total) * 100
        location = f'{assessment.barangay}, {assessment.municipality}'
        if match_pct == 100 and len(species_set) == len(db_set):
            if assessment.barangay_id == int(barangay_id) if barangay_id else False:
                warnings.append(
                    f'EXACT DUPLICATE: This {depth} file matches 100% ({len(species_set)} species) with assessment #{assessment.id} '
                    f'({location}, {assessment.assessment_date}).'
                )
            else:
                warnings.append(
                    f'EXACT DUPLICATE: This {depth} file matches 100% ({len(species_set)} species) with assessment #{assessment.id} '
                    f'at different location ({location}, {assessment.assessment_date}).'
                )
        elif match_pct > 75:
            warnings.append(
                f'High similarity ({match_pct:.0f}%) with assessment #{assessment.id} '
                f'({location}, {assessment.assessment_date}): {len(overlap)}/{total} species match exactly.'
            )
        elif match_pct > 50:
            warnings.append(
                f'{match_pct:.0f}% similarity with assessment #{assessment.id} '
                f'({location}, {assessment.assessment_date}): {len(overlap)}/{total} species match.'
            )

    return warnings


def validate_species_list(species_items):
    """
    Cross-check a list of (sub_category, major_category) tuples against the DB.
    Returns list of validation dicts.
    """
    import re
    existing_qs = Species.objects.all()
    existing_map = {(s.sub_category, s.major_category): s for s in existing_qs}

    results = []
    seen = set()
    for sub_cat, major_cat in species_items:
        key = (sub_cat, major_cat)
        if key in seen:
            continue
        seen.add(key)

        exact = existing_map.get(key)
        if exact:
            results.append({
                'sub_category': sub_cat,
                'major_category': major_cat,
                'status': 'existing',
                'matched_species_id': exact.id,
            })
        else:
            clean_name = re.sub(r'\s*\([^)]*\)\s*$', '', sub_cat).strip().lower()
            similar = None
            for s in existing_qs:
                db_clean = re.sub(r'\s*\([^)]*\)\s*$', '', s.sub_category).strip().lower()
                if db_clean == clean_name:
                    similar = s
                    break
            results.append({
                'sub_category': sub_cat,
                'major_category': major_cat,
                'status': 'new' if not similar else 'similar',
                'suggested_match': similar.id if similar else None,
                'suggested_name': str(similar) if similar else None,
            })
    return results


@login_required
@contributor_required
def upload_assessment(request):
    """Step 1: Assessment info + thesis PDF + images."""
    if request.method != 'POST':
        return render(request, 'contributor/upload_assessment.html', {
            'municipalities': Municipality.objects.all(),
            'custom_methodologies': CustomMethodology.objects.all(),
            'methodology_choices': Assessment.METHODOLOGY_CHOICES,
        })

    municipality_id = request.POST.get('municipality')
    barangay_id = request.POST.get('barangay')
    assessment_date = request.POST.get('assessment_date')
    methodology = request.POST.get('methodology', 'photo_quadrat')
    methodology_other = request.POST.get('methodology_other', '').strip()
    is_custom_methodology = False
    if methodology == 'other':
        if not methodology_other:
            messages.error(request, 'Please specify the methodology name.')
            return redirect('upload_assessment')
        methodology = methodology_other
        is_custom_methodology = True

    if not all([municipality_id, barangay_id, assessment_date]):
        messages.error(request, 'Please fill in all required fields.')
        return redirect('upload_assessment')

    # Save files to temp dir
    import tempfile, os, shutil
    tmp_dir = tempfile.mkdtemp()

    thesis_pdf = request.FILES.get('thesis_pdf')
    thesis_tmp = None
    if thesis_pdf:
        thesis_tmp = os.path.join(tmp_dir, 'thesis.pdf')
        with open(thesis_tmp, 'wb') as f:
            for chunk in thesis_pdf.chunks():
                f.write(chunk)

    images = request.FILES.getlist('images')
    image_tmps = []
    if images:
        img_dir = os.path.join(tmp_dir, 'images')
        os.makedirs(img_dir)
        for idx, img in enumerate(images):
            p = os.path.join(img_dir, f'image_{idx}{os.path.splitext(img.name)[1]}')
            with open(p, 'wb') as f:
                for chunk in img.chunks():
                    f.write(chunk)
            image_tmps.append(p)

    municipality = get_object_or_404(Municipality, id=municipality_id)
    barangay = get_object_or_404(Barangay, id=barangay_id, municipality=municipality)

    request.session['pending_assessment'] = {
        'municipality_id': municipality_id,
        'barangay_id': barangay_id,
        'municipality_name': str(municipality),
        'barangay_name': str(barangay),
        'assessment_date': assessment_date,
        'methodology': methodology,
        'is_custom_methodology': is_custom_methodology,
        'description': request.POST.get('description', '').strip(),
        'contributor_ids': [x.strip() for x in request.POST.get('contributors', '').split(',') if x.strip()],
        'tmp_dir': tmp_dir,
        'thesis_tmp': thesis_tmp,
        'image_tmps': image_tmps,
        'transects': [],
    }

    return redirect('add_transect')


@login_required
@contributor_required
def add_transect(request):
    """Step 2: Add transects. Each has shallow (lat/lng + excel) and deep (lat/lng + excel)."""
    pending = request.session.get('pending_assessment')
    if not pending:
        messages.error(request, 'Session expired. Please start again.')
        return redirect('upload_assessment')

    if request.method == 'POST':
        action = request.POST.get('action', 'add')

        if action == 'add':
            # Parse shallow excel
            shallow_file = request.FILES.get('shallow_excel')
            deep_file = request.FILES.get('deep_excel')
            shallow_start_lat = request.POST.get('shallow_start_lat')
            shallow_start_lng = request.POST.get('shallow_start_lng')
            shallow_end_lat = request.POST.get('shallow_end_lat')
            shallow_end_lng = request.POST.get('shallow_end_lng')
            deep_start_lat = request.POST.get('deep_start_lat')
            deep_start_lng = request.POST.get('deep_start_lng')
            deep_end_lat = request.POST.get('deep_end_lat')
            deep_end_lng = request.POST.get('deep_end_lng')

            if not shallow_file or not deep_file:
                messages.error(request, 'Both shallow and deep Excel files are required.')
                return redirect('add_transect')

            # Validate all 8 coordinate fields
            coord_fields = [
                ('shallow_start_lat', shallow_start_lat, -90, 90),
                ('shallow_start_lng', shallow_start_lng, -180, 180),
                ('shallow_end_lat', shallow_end_lat, -90, 90),
                ('shallow_end_lng', shallow_end_lng, -180, 180),
                ('deep_start_lat', deep_start_lat, -90, 90),
                ('deep_start_lng', deep_start_lng, -180, 180),
                ('deep_end_lat', deep_end_lat, -90, 90),
                ('deep_end_lng', deep_end_lng, -180, 180),
            ]
            coord_errors = []
            for field_name, val, min_val, max_val in coord_fields:
                label = field_name.replace('_', ' ').title()
                if not val or val.strip() == '':
                    coord_errors.append(f'{label} is required.')
                else:
                    try:
                        num = float(val)
                        if num < min_val or num > max_val:
                            coord_errors.append(f'{label} must be between {min_val} and {max_val}.')
                    except (TypeError, ValueError):
                        coord_errors.append(f'{label} must be a valid number.')
            if coord_errors:
                messages.error(request, 'Invalid coordinates. ' + ' '.join(coord_errors))
                return redirect('add_transect')

            import os
            tmp_dir = pending['tmp_dir']
            t_num = len(pending['transects']) + 1
            transect_info = {
                'transect_number': t_num,
                'shallow_start_lat': shallow_start_lat,
                'shallow_start_lng': shallow_start_lng,
                'shallow_end_lat': shallow_end_lat,
                'shallow_end_lng': shallow_end_lng,
                'deep_start_lat': deep_start_lat,
                'deep_start_lng': deep_start_lng,
                'deep_end_lat': deep_end_lat,
                'deep_end_lng': deep_end_lng,
                'shallow_species': [],
                'deep_species': [],
            }

            # Parse shallow excel
            ext = os.path.splitext(shallow_file.name)[1]
            shallow_path = os.path.join(tmp_dir, f't{t_num}_shallow{ext}')
            with open(shallow_path, 'wb') as f:
                for chunk in shallow_file.chunks():
                    f.write(chunk)
            shallow_errors = []
            try:
                species, parse_errors = parse_cpc_excel(shallow_path)
                shallow_errors = parse_errors
                if species:
                    transect_info['shallow_species'] = species
                    transect_info['shallow_filename'] = shallow_file.name
            except Exception as e:
                shallow_errors.append(str(e))

            # Parse deep excel
            ext = os.path.splitext(deep_file.name)[1]
            deep_path = os.path.join(tmp_dir, f't{t_num}_deep{ext}')
            with open(deep_path, 'wb') as f:
                for chunk in deep_file.chunks():
                    f.write(chunk)
            deep_errors = []
            try:
                species, parse_errors = parse_cpc_excel(deep_path)
                deep_errors = parse_errors
                if species:
                    transect_info['deep_species'] = species
                    transect_info['deep_filename'] = deep_file.name
            except Exception as e:
                deep_errors.append(str(e))

            # Single consolidated error if anything failed
            file_errors = []
            if not transect_info['shallow_species']:
                if shallow_errors:
                    file_errors.append(f"Shallow file: {'; '.join(shallow_errors)}")
                else:
                    file_errors.append("Shallow Excel has no valid species data")
            if not transect_info['deep_species']:
                if deep_errors:
                    file_errors.append(f"Deep file: {'; '.join(deep_errors)}")
                else:
                    file_errors.append("Deep Excel has no valid species data")

            if file_errors:
                messages.error(request, 'Invalid Excel file(s). ' + ' | '.join(file_errors) + '. Please check the format (Sub Category, Major Category, Mean).')
                return redirect('add_transect')

            # Check for duplicates
            dup_warnings = []
            if transect_info['shallow_species']:
                dup_warnings.extend(
                    check_duplicate_species(transect_info['shallow_species'], 'shallow', pending['transects'], barangay_id=pending.get('barangay_id'))
                )
            if transect_info['deep_species']:
                dup_warnings.extend(
                    check_duplicate_species(transect_info['deep_species'], 'deep', pending['transects'], barangay_id=pending.get('barangay_id'))
                )

            if dup_warnings:
                request.session['pending_dup_transect'] = transect_info
                request.session['dup_warnings'] = dup_warnings
                return redirect('confirm_duplicate')

            pending['transects'].append(transect_info)
            request.session['pending_assessment'] = pending
            if dup_warnings:
                request.session['dup_warnings_shown'] = True
            messages.success(request, f'Transect {t_num} added. Add another or proceed to preview.')
            return redirect('add_transect')

        elif action == 'remove':
            idx = int(request.POST.get('index', 0))
            if 0 <= idx < len(pending['transects']):
                pending['transects'].pop(idx)
                # Re-number
                for i, t in enumerate(pending['transects'], 1):
                    t['transect_number'] = i
                request.session['pending_assessment'] = pending
            return redirect('add_transect')

        elif action == 'preview':
            if not pending['transects']:
                messages.error(request, 'Add at least one transect before previewing.')
                return redirect('add_transect')
            return redirect('preview_assessment')

    # GET: show form with existing transects
    return render(request, 'contributor/add_transect.html', {
        'pending': pending,
        'transects': pending['transects'],
        'municipalities': Municipality.objects.all(),
    })


@login_required
@contributor_required
def confirm_duplicate(request):
    """Step 2b: Show duplicate warning and let user decide to proceed or cancel."""
    pending = request.session.get('pending_assessment')
    dup_transect = request.session.get('pending_dup_transect')
    dup_warnings = request.session.get('dup_warnings', [])

    if not pending or not dup_transect:
        messages.error(request, 'Session expired. Please start again.')
        return redirect('upload_assessment')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'proceed':
            pending['transects'].append(dup_transect)
            request.session['pending_assessment'] = pending
            del request.session['pending_dup_transect']
            del request.session['dup_warnings']
            messages.success(request, f'Transect {dup_transect["transect_number"]} added despite duplicate warnings.')
            return redirect('add_transect')
        else:
            del request.session['pending_dup_transect']
            del request.session['dup_warnings']
            messages.info(request, 'Transect discarded. No changes were made.')
            return redirect('add_transect')

    t = dup_transect
    species_count = len(t.get('shallow_species', [])) + len(t.get('deep_species', []))

    return render(request, 'contributor/confirm_duplicate.html', {
        'transect': t,
        'dup_warnings': dup_warnings,
        'species_count': species_count,
        'pending': pending,
    })


@login_required
@contributor_required
def preview_assessment(request):
    """Step 3: Show all parsed species, validate, let user confirm."""
    pending = request.session.get('pending_assessment')
    if not pending:
        messages.error(request, 'Session expired. Please start again.')
        return redirect('upload_assessment')

    # Collect all unique species across all transects and depths
    all_species = []
    for t in pending['transects']:
        for sp in t.get('shallow_species', []):
            all_species.append((sp['sub_category'], sp['major_category']))
        for sp in t.get('deep_species', []):
            all_species.append((sp['sub_category'], sp['major_category']))

    species_validation = validate_species_list(all_species)
    has_new = any(s['status'] in ('new', 'similar') for s in species_validation)
    user_role = request.user.profile.role if hasattr(request.user, 'profile') else 'contributor'

    contributor_ids = pending.get('contributor_ids', [])
    contributor_names = []
    if contributor_ids:
        contributor_names = list(
            Contributor.objects.filter(id__in=contributor_ids).values_list('first_name', flat=True)
        )
        # Re-fetch with full names
        contributor_names = [
            c.get_full_name() for c in Contributor.objects.filter(id__in=contributor_ids)
        ]

    return render(request, 'contributor/preview_assessment.html', {
        'pending': pending,
        'transects': pending['transects'],
        'species_validation': species_validation,
        'has_new_species': has_new,
        'user_role': user_role,
        'municipalities': Municipality.objects.all(),
        'contributor_names': contributor_names,
    })


@login_required
@contributor_required
def confirm_assessment(request):
    """Step 4: Save everything."""
    if request.method != 'POST':
        return redirect('upload_assessment')

    pending = request.session.get('pending_assessment')
    if not pending:
        messages.error(request, 'Session expired. Please start again.')
        return redirect('upload_assessment')

    municipality_id = pending['municipality_id']
    barangay_id = pending['barangay_id']
    assessment_date = pending['assessment_date']
    methodology = pending['methodology']

    municipality = get_object_or_404(Municipality, id=municipality_id)
    barangay = get_object_or_404(Barangay, id=barangay_id, municipality=municipality)

    # Auto-approve if admin/curator checked the box
    auto_approve = request.POST.get('auto_approve') == 'on'
    user_role = request.user.profile.role if hasattr(request.user, 'profile') else 'contributor'
    if auto_approve and user_role in ('admin', 'curator'):
        assessment_status = 'approved'
    else:
        assessment_status = 'submitted'

    assessment = Assessment.objects.create(
        municipality=municipality,
        barangay=barangay,
        assessment_date=assessment_date,
        methodology=methodology,
        status=assessment_status,
        uploaded_by=request.user,
        reviewed_by=request.user if auto_approve and user_role in ('admin', 'curator') else None,
        description=pending.get('description', ''),
    )

    # Add contributors
    contributor_ids = pending.get('contributor_ids', [])
    if contributor_ids:
        contributors = Contributor.objects.filter(id__in=contributor_ids)
        assessment.contributors.set(contributors)

    import shutil, os

    # Move thesis PDF
    thesis_tmp = pending.get('thesis_tmp')
    if thesis_tmp and os.path.exists(thesis_tmp):
        dest_dir = os.path.join(settings.MEDIA_ROOT, 'assessments', 'thesis', str(assessment.id))
        os.makedirs(dest_dir, exist_ok=True)
        dest = os.path.join(dest_dir, 'thesis.pdf')
        shutil.move(thesis_tmp, dest)
        assessment.thesis_pdf = os.path.join('assessments', 'thesis', str(assessment.id), 'thesis.pdf')
        assessment.save()

    # Move images
    for idx, img_tmp in enumerate(pending.get('image_tmps', [])):
        if os.path.exists(img_tmp):
            dest_dir = os.path.join(settings.MEDIA_ROOT, 'assessments', 'images', str(assessment.id))
            os.makedirs(dest_dir, exist_ok=True)
            ext = os.path.splitext(img_tmp)[1]
            dest = os.path.join(dest_dir, f'image_{idx}{ext}')
            shutil.move(img_tmp, dest)
            AssessmentImage.objects.create(assessment=assessment, image=os.path.join('assessments', 'images', str(assessment.id), f'image_{idx}{ext}'))

    # Create transects
    for t_info in pending['transects']:
        t_num = t_info['transect_number']

        def to_decimal(val):
            if val is None or val == '':
                return None
            try:
                d = Decimal(str(val))
                return d.quantize(Decimal('0.00000001'), rounding='ROUND_HALF_UP')
            except Exception:
                return None

        transect = Transect.objects.create(
            assessment=assessment,
            transect_number=t_num,
            shallow_start_lat=to_decimal(t_info.get('shallow_start_lat')),
            shallow_start_lng=to_decimal(t_info.get('shallow_start_lng')),
            shallow_end_lat=to_decimal(t_info.get('shallow_end_lat')),
            shallow_end_lng=to_decimal(t_info.get('shallow_end_lng')),
            deep_start_lat=to_decimal(t_info.get('deep_start_lat')),
            deep_start_lng=to_decimal(t_info.get('deep_start_lng')),
            deep_end_lat=to_decimal(t_info.get('deep_end_lat')),
            deep_end_lng=to_decimal(t_info.get('deep_end_lng')),
        )

        # Move shallow excel
        shallow_src = os.path.join(pending['tmp_dir'], f't{t_num}_shallow.xlsx')
        if not os.path.exists(shallow_src):
            shallow_src = os.path.join(pending['tmp_dir'], f't{t_num}_shallow.xls')
        if os.path.exists(shallow_src):
            dest_dir = os.path.join(settings.MEDIA_ROOT, 'assessments', 'transect_excel', str(assessment.id))
            os.makedirs(dest_dir, exist_ok=True)
            ext = os.path.splitext(shallow_src)[1]
            dest = os.path.join(dest_dir, f't{t_num}_shallow{ext}')
            shutil.move(shallow_src, dest)
            transect.shallow_excel = os.path.join('assessments', 'transect_excel', str(assessment.id), f't{t_num}_shallow{ext}')

        # Move deep excel
        deep_src = os.path.join(pending['tmp_dir'], f't{t_num}_deep.xlsx')
        if not os.path.exists(deep_src):
            deep_src = os.path.join(pending['tmp_dir'], f't{t_num}_deep.xls')
        if os.path.exists(deep_src):
            dest_dir = os.path.join(settings.MEDIA_ROOT, 'assessments', 'transect_excel', str(assessment.id))
            os.makedirs(dest_dir, exist_ok=True)
            ext = os.path.splitext(deep_src)[1]
            dest = os.path.join(dest_dir, f't{t_num}_deep{ext}')
            shutil.move(deep_src, dest)
            transect.deep_excel = os.path.join('assessments', 'transect_excel', str(assessment.id), f't{t_num}_deep{ext}')

        transect.save()

    # Cleanup
    tmp_dir = pending.get('tmp_dir')
    if tmp_dir and os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir, ignore_errors=True)
    del request.session['pending_assessment']

    if assessment_status == 'approved':
        species_count = populate_species_from_excel(assessment)
        save_barangay_transect_coords(assessment)
        if pending.get('is_custom_methodology'):
            CustomMethodology.objects.get_or_create(name=pending['methodology'])
        messages.success(request, f'Assessment auto-approved with {len(pending["transects"])} transect(s). {species_count} species record(s) created.')
    else:
        messages.success(request, f'Assessment submitted for review with {len(pending["transects"])} transect(s). Species will be recorded upon approval.')

    # Redirect to appropriate dashboard based on role
    if user_role == 'admin':
        return redirect('admin_dashboard')
    elif user_role == 'curator':
        return redirect('curator_dashboard')
    else:
        return redirect('contributor_dashboard')


@login_required
def get_transect_suggestions(request):
    """API: Get previous transect locations for a given barangay from BarangayTransect records."""
    barangay_id = request.GET.get('barangay_id')
    if not barangay_id:
        return JsonResponse({'suggestions': []})

    barangay_transects = BarangayTransect.objects.filter(
        barangay_id=barangay_id,
    ).order_by('pk')[:20]

    suggestions = []
    for idx, btc in enumerate(barangay_transects):
        suggestions.append({
            'id': btc.id,
            'ref_number': idx + 1,
            'shallow_start_lat': str(btc.shallow_start_lat) if btc.shallow_start_lat else None,
            'shallow_start_lng': str(btc.shallow_start_lng) if btc.shallow_start_lng else None,
            'shallow_end_lat': str(btc.shallow_end_lat) if btc.shallow_end_lat else None,
            'shallow_end_lng': str(btc.shallow_end_lng) if btc.shallow_end_lng else None,
            'deep_start_lat': str(btc.deep_start_lat) if btc.deep_start_lat else None,
            'deep_start_lng': str(btc.deep_start_lng) if btc.deep_start_lng else None,
            'deep_end_lat': str(btc.deep_end_lat) if btc.deep_end_lat else None,
            'deep_end_lng': str(btc.deep_end_lng) if btc.deep_end_lng else None,
        })

    return JsonResponse({'suggestions': suggestions})


@login_required
def check_reference_match(request):
    """API: Check if entered coordinates match any existing BarangayTransect reference."""
    barangay_id = request.GET.get('barangay_id')
    if not barangay_id:
        return JsonResponse({'match': None})

    def to_decimal(val):
        if not val:
            return None
        try:
            return Decimal(str(val)).quantize(Decimal('0.00000001'))
        except Exception:
            return None

    shallow_start_lat = to_decimal(request.GET.get('shallow_start_lat'))
    shallow_start_lng = to_decimal(request.GET.get('shallow_start_lng'))
    shallow_end_lat = to_decimal(request.GET.get('shallow_end_lat'))
    shallow_end_lng = to_decimal(request.GET.get('shallow_end_lng'))
    deep_start_lat = to_decimal(request.GET.get('deep_start_lat'))
    deep_start_lng = to_decimal(request.GET.get('deep_start_lng'))
    deep_end_lat = to_decimal(request.GET.get('deep_end_lat'))
    deep_end_lng = to_decimal(request.GET.get('deep_end_lng'))

    # Need at least 2 filled to attempt a match
    filled = sum(1 for v in [shallow_start_lat, shallow_start_lng, shallow_end_lat, shallow_end_lng,
                              deep_start_lat, deep_start_lng, deep_end_lat, deep_end_lng] if v is not None)
    if filled < 2:
        return JsonResponse({'match': None})

    match = BarangayTransect.objects.filter(
        barangay_id=barangay_id,
        shallow_start_lat=shallow_start_lat,
        shallow_start_lng=shallow_start_lng,
        shallow_end_lat=shallow_end_lat,
        shallow_end_lng=shallow_end_lng,
        deep_start_lat=deep_start_lat,
        deep_start_lng=deep_start_lng,
        deep_end_lat=deep_end_lat,
        deep_end_lng=deep_end_lng,
    ).first()

    if match:
        # Calculate ref_number by position
        refs = list(BarangayTransect.objects.filter(barangay_id=barangay_id).order_by('pk').values_list('pk', flat=True))
        ref_number = refs.index(match.pk) + 1 if match.pk in refs else '?'
        return JsonResponse({
            'match': True,
            'ref_number': ref_number,
            'ref_id': match.id,
            'source': match.source,
            'shallow_start_lat': str(match.shallow_start_lat) if match.shallow_start_lat else None,
            'shallow_start_lng': str(match.shallow_start_lng) if match.shallow_start_lng else None,
            'shallow_end_lat': str(match.shallow_end_lat) if match.shallow_end_lat else None,
            'shallow_end_lng': str(match.shallow_end_lng) if match.shallow_end_lng else None,
            'deep_start_lat': str(match.deep_start_lat) if match.deep_start_lat else None,
            'deep_start_lng': str(match.deep_start_lng) if match.deep_start_lng else None,
            'deep_end_lat': str(match.deep_end_lat) if match.deep_end_lat else None,
            'deep_end_lng': str(match.deep_end_lng) if match.deep_end_lng else None,
        })

    return JsonResponse({'match': None})


@login_required
def get_barangays(request):
    """API: Get barangays for a municipality."""
    municipality_id = request.GET.get('municipality_id')
    if not municipality_id:
        return JsonResponse({'barangays': []})
    barangays = Barangay.objects.filter(municipality_id=municipality_id).values('id', 'name')
    return JsonResponse({'barangays': list(barangays)})


@login_required
def search_contributors(request):
    """API: Search contributors by name (system users + external)."""
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})

    seen_ids = set()
    results = []

    # 1. Search approved system users — auto-link a Contributor if needed
    # Exclude the current user (uploader is already credited via uploaded_by)
    system_users = User.objects.filter(
        profile__status='approved',
    ).exclude(
        id=request.user.id
    ).filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(email__icontains=q)
    ).distinct()[:20]

    for u in system_users:
        # Find or create a Contributor linked to this user
        contrib, _ = Contributor.objects.get_or_create(
            user=u,
            defaults={
                'first_name': u.first_name or u.email.split('@')[0],
                'last_name': u.last_name or '',
            }
        )
        if contrib.id not in seen_ids:
            seen_ids.add(contrib.id)
            results.append({
                'id': contrib.id,
                'name': contrib.get_full_name(),
                'email': u.email,
                'is_user': True,
            })

    # 2. Search external contributors (no linked user)
    externals = Contributor.objects.filter(
        user__isnull=True,
    ).filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q)
    ).distinct()[:20]

    for c in externals:
        if c.id not in seen_ids:
            seen_ids.add(c.id)
            results.append({
                'id': c.id,
                'name': c.get_full_name(),
                'email': '',
                'is_user': False,
            })

    return JsonResponse({'results': results})


@login_required
def create_contributor(request):
    """API: Create a new external contributor (non-system user)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    first_name = data.get('first_name', '').strip()
    last_name = data.get('last_name', '').strip()
    middle_initial = data.get('middle_initial', '').strip()[:1].upper()
    suffix = data.get('suffix', '').strip()

    if not first_name or not last_name:
        return JsonResponse({'error': 'First and last name are required.'}, status=400)

    contributor = Contributor.objects.create(
        first_name=first_name,
        last_name=last_name,
        middle_initial=middle_initial,
        suffix=suffix,
        user=None,
    )

    return JsonResponse({
        'id': contributor.id,
        'name': contributor.get_full_name(),
        'is_user': False,
    })


@login_required
@login_required
@contributor_required
def my_assessments(request):
    """View list of contributor's own assessments."""
    assessments = Assessment.objects.filter(uploaded_by=request.user).order_by('-assessment_date')
    return render(request, 'contributor/my_assessments.html', {'assessments': assessments})


@login_required
@contributor_required
def contributor_assessment_detail(request, assessment_id):
    """Contributor: View their own assessment details."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects__species_data__species', 'images'),
        id=assessment_id,
        uploaded_by=request.user,
    )
    transects = assessment.transects.prefetch_related('species_data__species').all()

    parsed_species = []
    has_species_records = TransectSpecies.objects.filter(transect__assessment=assessment).exists()
    if not has_species_records:
        for transect in transects:
            if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
                sp, _ = parse_cpc_excel(transect.shallow_excel.path)
                for s in sp:
                    s['depth'] = 'Shallow'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)
            if transect.deep_excel and os.path.exists(transect.deep_excel.path):
                sp, _ = parse_cpc_excel(transect.deep_excel.path)
                for s in sp:
                    s['depth'] = 'Deep'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)

    return render(request, 'contributor/assessment_detail.html', {
        'assessment': assessment,
        'transects': transects,
        'parsed_species': parsed_species,
        'has_species_records': has_species_records,
    })


@login_required
@contributor_required
def delete_assessment(request, assessment_id):
    """Delete assessment. Curators cannot delete. Admin can delete rejected only. Contributors can delete submitted/rejected."""
    if request.method != 'POST':
        return redirect('my_assessments')

    user_role = request.user.profile.role if hasattr(request.user, 'profile') else 'contributor'

    # Curators cannot delete assessments
    if user_role == 'curator':
        messages.error(request, 'Curators do not have permission to delete assessments.')
        return redirect('curator_assessments')

    with transaction.atomic():
        if user_role == 'admin':
            assessment = get_object_or_404(
                Assessment.objects.select_for_update(), id=assessment_id
            )
        else:
            assessment = get_object_or_404(
                Assessment.objects.select_for_update(), id=assessment_id, uploaded_by=request.user
            )

        # Admin can only delete rejected; contributors can delete submitted or rejected
        if user_role == 'admin' and assessment.status != 'rejected':
            messages.error(request, 'Admin can only delete rejected assessments.')
            return redirect('admin_assessments')

        if user_role == 'contributor' and assessment.status not in ('submitted', 'rejected'):
            messages.error(request, 'You can only delete assessments that are pending or rejected.')
            return redirect('my_assessments')

        # Delete associated files
        if assessment.thesis_pdf:
            thesis_path = assessment.thesis_pdf.path if hasattr(assessment.thesis_pdf, 'path') else None
            if thesis_path and os.path.exists(thesis_path):
                os.remove(thesis_path)
        for img in assessment.images.all():
            img_path = img.image.path if hasattr(img.image, 'path') else None
            if img_path and os.path.exists(img_path):
                os.remove(img_path)
        for transect in assessment.transects.all():
            if transect.shallow_excel:
                path = transect.shallow_excel.path if hasattr(transect.shallow_excel, 'path') else None
                if path and os.path.exists(path):
                    os.remove(path)
            if transect.deep_excel:
                path = transect.deep_excel.path if hasattr(transect.deep_excel, 'path') else None
                if path and os.path.exists(path):
                    os.remove(path)

        assessment.delete()
        messages.success(request, f'Assessment #{assessment_id} has been deleted.')
    if user_role == 'admin':
        return redirect('admin_assessments')
    return redirect('my_assessments')


# ==================== ADMIN ASSESSMENT REVIEW VIEWS ====================

@login_required
@admin_required
def admin_assessments(request):
    """Admin: List all assessments with filters."""
    assessments = Assessment.objects.select_related(
        'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
    ).prefetch_related('transects').all()

    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    if status_filter:
        assessments = assessments.filter(status=status_filter)
    if search_query:
        assessments = assessments.filter(
            Q(barangay__name__icontains=search_query) |
            Q(municipality__name__icontains=search_query) |
            Q(uploaded_by__username__icontains=search_query)
        )

    stats = {
        'total': Assessment.objects.count(),
        'submitted': Assessment.objects.filter(status='submitted').count(),
        'approved': Assessment.objects.filter(status='approved').count(),
        'rejected': Assessment.objects.filter(status='rejected').count(),
        'draft': Assessment.objects.filter(status='draft').count(),
    }

    context = {
        'assessments': assessments,
        'status_filter': status_filter,
        'search_query': search_query,
        'stats': stats,
    }
    return render(request, 'admin/assessments/index.html', context)


@login_required
@admin_required
def admin_bulk_delete_assessments(request):
    """Admin: Bulk delete selected assessments. Only rejected and draft can be deleted."""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.warning(request, 'No assessments selected.')
            return redirect('admin_assessments')
        deleted = []
        skipped = []
        for aid in ids:
            try:
                a = Assessment.objects.get(id=aid)
            except Assessment.DoesNotExist:
                continue
            if a.status in ('submitted', 'approved'):
                skipped.append(f'#{a.id} ({a.get_status_display()})')
            else:
                deleted.append(f'#{a.id} ({a.get_status_display()})')
                a.delete()
        if deleted:
            messages.success(request, f'Deleted {len(deleted)} assessment(s): {", ".join(deleted)}.')
        if skipped:
            messages.warning(request, f'Skipped {len(skipped)} (pending or approved): {", ".join(skipped)}.')
    return redirect('admin_assessments')


@login_required
@admin_required
def admin_assessment_detail(request, assessment_id):
    """Admin: View assessment details with duplicate detection."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects__species_data__species', 'images'),
        id=assessment_id,
    )
    transects = assessment.transects.prefetch_related('species_data__species').all()

    # Parse Excel files on-the-fly if no species records yet
    parsed_species = []
    has_species_records = TransectSpecies.objects.filter(transect__assessment=assessment).exists()
    if not has_species_records:
        for transect in transects:
            if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
                sp, _ = parse_cpc_excel(transect.shallow_excel.path)
                for s in sp:
                    s['depth'] = 'Shallow'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)
            if transect.deep_excel and os.path.exists(transect.deep_excel.path):
                sp, _ = parse_cpc_excel(transect.deep_excel.path)
                for s in sp:
                    s['depth'] = 'Deep'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)

    # Find potential duplicate assessments
    similar_assessments = []
    current_species = set()

    # Use parsed species if no DB records
    if not has_species_records:
        for sp in parsed_species:
                current_species.add((sp['sub_category'], sp['major_category'], round(sp['cover'], 2)))
    else:
        for ts in TransectSpecies.objects.filter(
            transect__assessment=assessment
        ).select_related('species'):
            current_species.add((ts.species.sub_category, ts.species.major_category, float(ts.cover)))

    if current_species:
        other_assessments = Assessment.objects.filter(
            barangay=assessment.barangay,
            status='approved',
        ).exclude(id=assessment.id).order_by('-assessment_date')[:10]

        for other in other_assessments:
            other_species = set()
            for ts in TransectSpecies.objects.filter(
                transect__assessment=other
            ).select_related('species'):
                other_species.add((ts.species.sub_category, ts.species.major_category, float(ts.cover)))

            if not other_species:
                continue

            overlap = current_species & other_species
            total = max(len(current_species), len(other_species))
            if total == 0:
                continue

            match_pct = (len(overlap) / total) * 100
            if match_pct >= 50:
                similar_assessments.append({
                    'assessment': other,
                    'match_count': len(overlap),
                    'total': total,
                    'match_pct': round(match_pct, 1),
                    'is_exact': match_pct == 100 and len(current_species) == len(other_species),
                })

    return render(request, 'admin/assessments/detail.html', {
        'assessment': assessment,
        'transects': transects,
        'similar_assessments': similar_assessments,
        'parsed_species': parsed_species,
        'has_species_records': has_species_records,
    })


@login_required
@admin_required
def admin_confirm_approval(request, assessment_id):
    """Admin: Show approval warnings before confirming."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by'
        ),
        id=assessment_id,
        status='submitted',
    )
    transects = assessment.transects.all()

    # Parse Excel files to get species data
    all_species = []
    for transect in transects:
        if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
            species_list, _ = parse_cpc_excel(transect.shallow_excel.path)
            for sp in species_list:
                sp['transect'] = transect.transect_number
                sp['depth'] = 'Shallow'
                all_species.append(sp)
        if transect.deep_excel and os.path.exists(transect.deep_excel.path):
            species_list, _ = parse_cpc_excel(transect.deep_excel.path)
            for sp in species_list:
                sp['transect'] = transect.transect_number
                sp['depth'] = 'Deep'
                all_species.append(sp)

    # Check for duplicates against approved assessments
    current_set = {(sp['sub_category'], sp['major_category'], round(sp['cover'], 2)) for sp in all_species}
    duplicate_warnings = []
    if current_set:
        approved = Assessment.objects.filter(
            barangay=assessment.barangay, status='approved'
        ).exclude(id=assessment.id).order_by('-assessment_date')[:10]
        for other in approved:
            other_species = TransectSpecies.objects.filter(
                transect__assessment=other
            ).select_related('species')
            other_set = {(ts.species.sub_category, ts.species.major_category, round(float(ts.cover), 2)) for ts in other_species}
            if not other_set:
                continue
            overlap = current_set & other_set
            total = max(len(current_set), len(other_set))
            if total == 0:
                continue
            match_pct = (len(overlap) / total) * 100
            if match_pct >= 50:
                duplicate_warnings.append({
                    'assessment': other,
                    'match_count': len(overlap),
                    'total': total,
                    'match_pct': round(match_pct, 1),
                    'is_exact': match_pct == 100 and len(current_set) == len(other_set),
                    'overlap_species': list(overlap)[:5],
                })

    # Check for new species not in the Species table
    new_species = []
    new_keys = set()
    for sp in all_species:
        exists = Species.objects.filter(
            sub_category__iexact=sp['sub_category'],
            major_category__iexact=sp['major_category'],
        ).exists()
        sp['is_new'] = not exists
        if not exists:
            new_species.append(sp)
            new_keys.add((sp['sub_category'].lower(), sp['major_category'].lower()))

    # Unique new species
    seen = set()
    unique_new = []
    for ns in new_species:
        key = (ns['sub_category'].lower(), ns['major_category'].lower())
        if key not in seen:
            seen.add(key)
            unique_new.append(ns)

    has_warnings = bool(duplicate_warnings or unique_new)

    return render(request, 'admin/assessments/confirm_approval.html', {
        'assessment': assessment,
        'transects': transects,
        'all_species': all_species,
        'duplicate_warnings': duplicate_warnings,
        'new_species': unique_new,
        'has_warnings': has_warnings,
    })


@login_required
@admin_required
def admin_assessment_action(request, assessment_id):
    """Admin: Approve, reject, or return assessment to pending."""
    with transaction.atomic():
        assessment = get_object_or_404(
            Assessment.objects.select_for_update(), id=assessment_id
        )

        action = request.POST.get('action')
        built_in = {val for val, label in Assessment.METHODOLOGY_CHOICES}

        if action == 'approve':
            assessment.status = 'approved'
            assessment.reviewed_by = request.user
            assessment.approved_at = timezone.now()
            assessment.save()
            species_count = populate_species_from_excel(assessment)
            save_barangay_transect_coords(assessment)
            if assessment.methodology not in built_in:
                CustomMethodology.objects.get_or_create(name=assessment.methodology)
            messages.success(request, f'Assessment #{assessment.id} approved. {species_count} species record(s) created.')
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '')
            assessment.status = 'rejected'
            assessment.reviewed_by = request.user
            assessment.approved_at = None
            assessment.notes = rejection_reason
            assessment.save()
            messages.info(request, f'Assessment #{assessment.id} has been rejected.')
        elif action == 'return_to_pending':
            assessment.status = 'submitted'
            assessment.reviewed_by = None
            assessment.approved_at = None
            assessment.notes = ''
            assessment.save()
            if assessment.methodology not in built_in:
                other_approved = Assessment.objects.filter(
                    status='approved', methodology=assessment.methodology
                ).exclude(id=assessment.id).exists()
                if not other_approved:
                    CustomMethodology.objects.filter(name=assessment.methodology).delete()
            cleanup_barangay_transect_coords(assessment)
            transects = assessment.transects.all()
            species_ids = list(TransectSpecies.objects.filter(
                transect__in=transects
            ).values_list('species_id', flat=True).distinct())
            deleted, _ = TransectSpecies.objects.filter(transect__in=transects).delete()
            if species_ids:
                other_approved_species = set(TransectSpecies.objects.filter(
                    transect__assessment__status='approved'
                ).exclude(transect__assessment=assessment).values_list('species_id', flat=True).distinct())
                orphan_ids = set(species_ids) - other_approved_species
                if orphan_ids:
                    Species.objects.filter(id__in=orphan_ids, source=Species.SOURCE_ASSESSMENT).delete()
            messages.success(request, f'Assessment #{assessment.id} returned to pending. {deleted} species record(s) removed.')
        else:
            messages.error(request, 'Invalid action.')

    return redirect('admin_assessments')


# ==================== ADMIN TRANSECT COORDINATE MANAGEMENT VIEWS ====================

@login_required
@admin_required
def admin_manage_transect_coords(request):
    """Admin: List barangays with transect coordinate counts."""
    barangays = Barangay.objects.select_related('municipality').annotate(
        coord_count=Count('barangay_transects'),
    ).order_by('municipality__name', 'name')
    return render(request, 'admin/transect_coords/index.html', {
        'barangays': barangays,
    })


@login_required
@admin_required
def admin_barangay_transect_coords(request, barangay_id):
    """Admin: View and manage transect coordinates for a specific barangay."""
    barangay = get_object_or_404(Barangay, id=barangay_id)
    transect_coords = BarangayTransect.objects.filter(barangay=barangay).order_by('pk')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            def to_decimal(val):
                val = val.strip()
                if val == '':
                    return None
                from decimal import Decimal, InvalidOperation
                try:
                    return Decimal(val).quantize(Decimal('0.00000001'))
                except InvalidOperation:
                    return None

            fields = {
                'shallow_start_lat': request.POST.get('shallow_start_lat', ''),
                'shallow_start_lng': request.POST.get('shallow_start_lng', ''),
                'shallow_end_lat': request.POST.get('shallow_end_lat', ''),
                'shallow_end_lng': request.POST.get('shallow_end_lng', ''),
                'deep_start_lat': request.POST.get('deep_start_lat', ''),
                'deep_start_lng': request.POST.get('deep_start_lng', ''),
                'deep_end_lat': request.POST.get('deep_end_lat', ''),
                'deep_end_lng': request.POST.get('deep_end_lng', ''),
            }
            coord_errors = []
            for fname, val in fields.items():
                label = fname.replace('_', ' ').title()
                val = val.strip()
                if val == '':
                    coord_errors.append(f'{label} is required.')
                else:
                    try:
                        num = float(val)
                        if fname.endswith('_lat') and (num < -90 or num > 90):
                            coord_errors.append(f'{label} must be between -90 and 90.')
                        elif fname.endswith('_lng') and (num < -180 or num > 180):
                            coord_errors.append(f'{label} must be between -180 and 180.')
                    except (TypeError, ValueError):
                        coord_errors.append(f'{label} must be a valid number.')
            if coord_errors:
                messages.error(request, 'Invalid coordinates. ' + ' '.join(coord_errors))
                return redirect('admin_barangay_transect_coords', barangay_id=barangay.id)

            BarangayTransect.objects.create(
                barangay=barangay,
                shallow_start_lat=to_decimal(fields['shallow_start_lat']),
                shallow_start_lng=to_decimal(fields['shallow_start_lng']),
                shallow_end_lat=to_decimal(fields['shallow_end_lat']),
                shallow_end_lng=to_decimal(fields['shallow_end_lng']),
                deep_start_lat=to_decimal(fields['deep_start_lat']),
                deep_start_lng=to_decimal(fields['deep_start_lng']),
                deep_end_lat=to_decimal(fields['deep_end_lat']),
                deep_end_lng=to_decimal(fields['deep_end_lng']),
                source='manual',
            )
            messages.success(request, f'Reference coordinates added for {barangay.name}.')
            return redirect('admin_barangay_transect_coords', barangay_id=barangay.id)

        elif action == 'edit':
            btc_id = request.POST.get('transect_id')
            btc = get_object_or_404(BarangayTransect, id=btc_id, barangay=barangay)

            fields = {
                'shallow_start_lat': request.POST.get('shallow_start_lat', ''),
                'shallow_start_lng': request.POST.get('shallow_start_lng', ''),
                'shallow_end_lat': request.POST.get('shallow_end_lat', ''),
                'shallow_end_lng': request.POST.get('shallow_end_lng', ''),
                'deep_start_lat': request.POST.get('deep_start_lat', ''),
                'deep_start_lng': request.POST.get('deep_start_lng', ''),
                'deep_end_lat': request.POST.get('deep_end_lat', ''),
                'deep_end_lng': request.POST.get('deep_end_lng', ''),
            }
            coord_errors = []
            for fname, val in fields.items():
                label = fname.replace('_', ' ').title()
                val = val.strip()
                if val == '':
                    coord_errors.append(f'{label} is required.')
                else:
                    try:
                        num = float(val)
                        if fname.endswith('_lat') and (num < -90 or num > 90):
                            coord_errors.append(f'{label} must be between -90 and 90.')
                        elif fname.endswith('_lng') and (num < -180 or num > 180):
                            coord_errors.append(f'{label} must be between -180 and 180.')
                    except (TypeError, ValueError):
                        coord_errors.append(f'{label} must be a valid number.')
            if coord_errors:
                messages.error(request, 'Invalid coordinates. ' + ' '.join(coord_errors))
                return redirect('admin_barangay_transect_coords', barangay_id=barangay.id)

            def to_decimal(val):
                val = val.strip()
                if val == '':
                    return None
                from decimal import Decimal, InvalidOperation
                try:
                    return Decimal(val).quantize(Decimal('0.00000001'))
                except InvalidOperation:
                    return None

            btc.shallow_start_lat = to_decimal(fields['shallow_start_lat'])
            btc.shallow_start_lng = to_decimal(fields['shallow_start_lng'])
            btc.shallow_end_lat = to_decimal(fields['shallow_end_lat'])
            btc.shallow_end_lng = to_decimal(fields['shallow_end_lng'])
            btc.deep_start_lat = to_decimal(fields['deep_start_lat'])
            btc.deep_start_lng = to_decimal(fields['deep_start_lng'])
            btc.deep_end_lat = to_decimal(fields['deep_end_lat'])
            btc.deep_end_lng = to_decimal(fields['deep_end_lng'])
            btc.save()
            messages.success(request, f'Reference coordinates updated for {barangay.name}.')
            return redirect('admin_barangay_transect_coords', barangay_id=barangay.id)

        elif action == 'delete':
            btc_id = request.POST.get('transect_id')
            btc = get_object_or_404(BarangayTransect, id=btc_id, barangay=barangay)
            if btc.source == 'assessment':
                messages.error(request, 'This reference is from an approved assessment and cannot be deleted.')
            else:
                btc.delete()
                messages.success(request, f'Transect {btc.transect_number} deleted from {barangay.name}.')
            return redirect('admin_barangay_transect_coords', barangay_id=barangay.id)

    return render(request, 'admin/transect_coords/barangay.html', {
        'barangay': barangay,
        'transect_coords': transect_coords,
    })


# ==================== ADMIN LOCATION MANAGEMENT VIEWS ====================

@login_required
@admin_required
def admin_manage_locations(request):
    """Admin: List all municipalities with barangay counts."""
    municipalities = Municipality.objects.annotate(
        barangay_count=Count('barangays'),
        approved_assessment_count=Count('barangays__assessments', filter=Q(barangays__assessments__status='approved')),
        total_assessment_count=Count('barangays__assessments'),
    ).order_by('name')
    return render(request, 'admin/locations/index.html', {
        'municipalities': municipalities,
    })


@login_required
@admin_required
def admin_add_municipality(request):
    """Admin: Add a new municipality."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Municipality name is required.')
            return redirect('admin_manage_locations')
        if Municipality.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Municipality "{name}" already exists.')
            return redirect('admin_manage_locations')
        Municipality.objects.create(name=name)
        messages.success(request, f'Municipality "{name}" added successfully.')
        return redirect('admin_manage_locations')
    return redirect('admin_manage_locations')


@login_required
@admin_required
def admin_edit_municipality(request, municipality_id):
    """Admin: Edit a municipality."""
    municipality = get_object_or_404(Municipality, id=municipality_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Municipality name is required.')
            return redirect('admin_manage_locations')
        if Municipality.objects.filter(name__iexact=name).exclude(id=municipality_id).exists():
            messages.error(request, f'Municipality "{name}" already exists.')
            return redirect('admin_manage_locations')
        municipality.name = name
        municipality.save()
        messages.success(request, f'Municipality updated to "{name}".')
        return redirect('admin_manage_locations')
    return redirect('admin_manage_locations')


@login_required
@admin_required
def admin_delete_municipality(request, municipality_id):
    """Admin: Delete a municipality."""
    municipality = get_object_or_404(Municipality, id=municipality_id)
    if request.method == 'POST':
        name = municipality.name
        if municipality.assessments.exists():
            messages.error(request, f'Cannot delete "{name}" — it is used in existing assessments.')
            return redirect('admin_manage_locations')
        municipality.delete()
        messages.success(request, f'Municipality "{name}" deleted.')
    return redirect('admin_manage_locations')


@login_required
@admin_required
def admin_bulk_delete_municipalities(request):
    """Admin: Bulk delete selected municipalities (skips those with any assessments)."""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.warning(request, 'No municipalities selected.')
            return redirect('admin_manage_locations')
        deleted = []
        skipped = []
        for mid in ids:
            try:
                m = Municipality.objects.get(id=mid)
            except Municipality.DoesNotExist:
                continue
            if m.assessments.exists():
                skipped.append(m.name)
            else:
                deleted.append(m.name)
                m.delete()
        if deleted:
            messages.success(request, f'Deleted {len(deleted)} municipality: {", ".join(deleted)}.')
        if skipped:
            messages.warning(request, f'Skipped {len(skipped)} (used in assessments): {", ".join(skipped)}.')
    return redirect('admin_manage_locations')


@login_required
@admin_required
def admin_manage_barangays(request, municipality_id):
    """Admin: List and manage barangays for a municipality, with transect coordinate management."""
    municipality = get_object_or_404(Municipality, id=municipality_id)
    barangays = municipality.barangays.annotate(
        coord_count=Count('barangay_transects'),
        approved_assessment_count=Count('assessments', filter=Q(assessments__status='approved')),
        total_assessment_count=Count('assessments'),
    ).order_by('name')

    # Pre-fetch transect coords for each barangay
    all_transects = list(BarangayTransect.objects.filter(
        barangay__municipality=municipality
    ).order_by('pk'))
    transect_coords = {}
    for btc in all_transects:
        transect_coords.setdefault(btc.barangay_id, []).append(btc)

    # Build per-barangay list for template (avoids dict key lookup issue)
    barangay_data = []
    for b in barangays:
        barangay_data.append({
            'barangay': b,
            'coords': transect_coords.get(b.id, []),
        })

    # Build JSON for JS map init
    transect_data_json = {}
    for btc in all_transects:
        bid = str(btc.barangay_id)
        transect_data_json.setdefault(bid, []).append({
            'id': btc.id,
            'sLat': float(btc.shallow_start_lat) if btc.shallow_start_lat else None,
            'sLng': float(btc.shallow_start_lng) if btc.shallow_start_lng else None,
            'eLat': float(btc.shallow_end_lat) if btc.shallow_end_lat else None,
            'eLng': float(btc.shallow_end_lng) if btc.shallow_end_lng else None,
            'dSLat': float(btc.deep_start_lat) if btc.deep_start_lat else None,
            'dSLng': float(btc.deep_start_lng) if btc.deep_start_lng else None,
            'dELat': float(btc.deep_end_lat) if btc.deep_end_lat else None,
            'dELng': float(btc.deep_end_lng) if btc.deep_end_lng else None,
        })

    # Handle transect coordinate POST actions
    if request.method == 'POST':
        post_action = request.POST.get('transect_action')
        brgy_id = request.POST.get('barangay_id')

        if post_action == 'add_transect':
            barangay = get_object_or_404(Barangay, id=brgy_id, municipality=municipality)

            def to_decimal(val):
                val = (val or '').strip()
                if val == '':
                    return None
                from decimal import Decimal, InvalidOperation
                try:
                    return Decimal(val).quantize(Decimal('0.00000001'))
                except InvalidOperation:
                    return None

            fields = {
                'shallow_start_lat': request.POST.get('shallow_start_lat', ''),
                'shallow_start_lng': request.POST.get('shallow_start_lng', ''),
                'shallow_end_lat': request.POST.get('shallow_end_lat', ''),
                'shallow_end_lng': request.POST.get('shallow_end_lng', ''),
                'deep_start_lat': request.POST.get('deep_start_lat', ''),
                'deep_start_lng': request.POST.get('deep_start_lng', ''),
                'deep_end_lat': request.POST.get('deep_end_lat', ''),
                'deep_end_lng': request.POST.get('deep_end_lng', ''),
            }
            coord_errors = []
            for fname, val in fields.items():
                label = fname.replace('_', ' ').title()
                val = val.strip()
                if val == '':
                    coord_errors.append(f'{label} is required.')
                else:
                    try:
                        num = float(val)
                        if fname.endswith('_lat') and (num < -90 or num > 90):
                            coord_errors.append(f'{label} must be between -90 and 90.')
                        elif fname.endswith('_lng') and (num < -180 or num > 180):
                            coord_errors.append(f'{label} must be between -180 and 180.')
                    except (TypeError, ValueError):
                        coord_errors.append(f'{label} must be a valid number.')
            if coord_errors:
                messages.error(request, 'Invalid coordinates. ' + ' '.join(coord_errors))
                return redirect('admin_manage_barangays', municipality_id=municipality_id)

            BarangayTransect.objects.create(
                barangay=barangay,
                shallow_start_lat=to_decimal(fields['shallow_start_lat']),
                shallow_start_lng=to_decimal(fields['shallow_start_lng']),
                shallow_end_lat=to_decimal(fields['shallow_end_lat']),
                shallow_end_lng=to_decimal(fields['shallow_end_lng']),
                deep_start_lat=to_decimal(fields['deep_start_lat']),
                deep_start_lng=to_decimal(fields['deep_start_lng']),
                deep_end_lat=to_decimal(fields['deep_end_lat']),
                deep_end_lng=to_decimal(fields['deep_end_lng']),
                source='manual',
            )
            messages.success(request, f'Reference coordinates added for {barangay.name}.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)

        elif post_action == 'edit_transect':
            btc_id = request.POST.get('transect_id')
            btc = get_object_or_404(BarangayTransect, id=btc_id, barangay__municipality=municipality)
            if btc.source == 'assessment':
                messages.error(request, 'This reference is from an approved assessment and cannot be edited.')
                return redirect('admin_manage_barangays', municipality_id=municipality_id)

            def to_decimal(val):
                val = (val or '').strip()
                if val == '':
                    return None
                from decimal import Decimal, InvalidOperation
                try:
                    return Decimal(val).quantize(Decimal('0.00000001'))
                except InvalidOperation:
                    return None

            fields = {
                'shallow_start_lat': request.POST.get('shallow_start_lat', ''),
                'shallow_start_lng': request.POST.get('shallow_start_lng', ''),
                'shallow_end_lat': request.POST.get('shallow_end_lat', ''),
                'shallow_end_lng': request.POST.get('shallow_end_lng', ''),
                'deep_start_lat': request.POST.get('deep_start_lat', ''),
                'deep_start_lng': request.POST.get('deep_start_lng', ''),
                'deep_end_lat': request.POST.get('deep_end_lat', ''),
                'deep_end_lng': request.POST.get('deep_end_lng', ''),
            }
            coord_errors = []
            for fname, val in fields.items():
                label = fname.replace('_', ' ').title()
                val = val.strip()
                if val == '':
                    coord_errors.append(f'{label} is required.')
                else:
                    try:
                        num = float(val)
                        if fname.endswith('_lat') and (num < -90 or num > 90):
                            coord_errors.append(f'{label} must be between -90 and 90.')
                        elif fname.endswith('_lng') and (num < -180 or num > 180):
                            coord_errors.append(f'{label} must be between -180 and 180.')
                    except (TypeError, ValueError):
                        coord_errors.append(f'{label} must be a valid number.')
            if coord_errors:
                messages.error(request, 'Invalid coordinates. ' + ' '.join(coord_errors))
                return redirect('admin_manage_barangays', municipality_id=municipality_id)

            btc.shallow_start_lat = to_decimal(fields['shallow_start_lat'])
            btc.shallow_start_lng = to_decimal(fields['shallow_start_lng'])
            btc.shallow_end_lat = to_decimal(fields['shallow_end_lat'])
            btc.shallow_end_lng = to_decimal(fields['shallow_end_lng'])
            btc.deep_start_lat = to_decimal(fields['deep_start_lat'])
            btc.deep_start_lng = to_decimal(fields['deep_start_lng'])
            btc.deep_end_lat = to_decimal(fields['deep_end_lat'])
            btc.deep_end_lng = to_decimal(fields['deep_end_lng'])
            btc.save()
            messages.success(request, f'Reference coordinates updated for {btc.barangay.name}.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)

        elif post_action == 'delete_transect':
            btc_id = request.POST.get('transect_id')
            btc = get_object_or_404(BarangayTransect, id=btc_id, barangay__municipality=municipality)
            if btc.source == 'assessment':
                messages.error(request, 'This reference is from an approved assessment and cannot be deleted.')
                return redirect('admin_manage_barangays', municipality_id=municipality_id)
            btc.delete()
            messages.success(request, f'Reference coordinates deleted from {btc.barangay.name}.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)

    return render(request, 'admin/locations/barangays.html', {
        'municipality': municipality,
        'barangays': barangays,
        'barangay_data': barangay_data,
        'all_transects': all_transects,
        'transect_data_json': json.dumps(transect_data_json),
    })


@login_required
@admin_required
def admin_add_barangay(request, municipality_id):
    """Admin: Add a barangay to a municipality."""
    municipality = get_object_or_404(Municipality, id=municipality_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Barangay name is required.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        if Barangay.objects.filter(name__iexact=name, municipality=municipality).exists():
            messages.error(request, f'Barangay "{name}" already exists in {municipality.name}.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        Barangay.objects.create(name=name, municipality=municipality)
        messages.success(request, f'Barangay "{name}" added to {municipality.name}.')
    return redirect('admin_manage_barangays', municipality_id=municipality_id)


@login_required
@admin_required
def admin_edit_barangay(request, barangay_id):
    """Admin: Edit a barangay."""
    barangay = get_object_or_404(Barangay, id=barangay_id)
    municipality_id = barangay.municipality_id
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Barangay name is required.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        if Barangay.objects.filter(name__iexact=name, municipality=barangay.municipality).exclude(id=barangay_id).exists():
            messages.error(request, f'Barangay "{name}" already exists in {barangay.municipality.name}.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        barangay.name = name
        barangay.save()
        messages.success(request, f'Barangay updated to "{name}".')
    return redirect('admin_manage_barangays', municipality_id=municipality_id)


@login_required
@admin_required
def admin_delete_barangay(request, barangay_id):
    """Admin: Delete a barangay."""
    barangay = get_object_or_404(Barangay, id=barangay_id)
    municipality_id = barangay.municipality_id
    if request.method == 'POST':
        name = barangay.name
        if barangay.assessments.exists():
            messages.error(request, f'Cannot delete "{name}" — it is used in existing assessments.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        barangay.delete()
        messages.success(request, f'Barangay "{name}" deleted.')
    return redirect('admin_manage_barangays', municipality_id=municipality_id)


@login_required
@admin_required
def admin_bulk_delete_barangays(request, municipality_id):
    """Admin: Bulk delete selected barangays (skips those with any assessments)."""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.warning(request, 'No barangays selected.')
            return redirect('admin_manage_barangays', municipality_id=municipality_id)
        deleted = []
        skipped = []
        for bid in ids:
            try:
                b = Barangay.objects.get(id=bid, municipality_id=municipality_id)
            except Barangay.DoesNotExist:
                continue
            if b.assessments.exists():
                skipped.append(b.name)
            else:
                deleted.append(b.name)
                b.delete()
        if deleted:
            messages.success(request, f'Deleted {len(deleted)} barangay(s): {", ".join(deleted)}.')
        if skipped:
            messages.warning(request, f'Skipped {len(skipped)} (used in assessments): {", ".join(skipped)}.')
    return redirect('admin_manage_barangays', municipality_id=municipality_id)


# ==================== ADMIN SPECIES MANAGEMENT VIEWS ====================

@login_required
@admin_required
def admin_manage_species(request):
    """Admin: List all families (major categories) with species counts."""
    families = (
        Species.objects.values('major_category')
        .annotate(species_count=Count('id'))
        .order_by('major_category')
    )
    return render(request, 'admin/species/index.html', {
        'families': families,
        'total_species': Species.objects.count(),
    })


@login_required
@admin_required
def admin_add_family(request):
    """Admin: Add a new family (major category)."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Family name is required.')
            return redirect('admin_manage_species')
        if Species.objects.filter(major_category__iexact=name).exists():
            messages.error(request, f'Family "{name}" already exists.')
            return redirect('admin_manage_species')
        code = request.POST.get('code', '').strip()
        sub = request.POST.get('sub_category', '').strip()
        if sub:
            Species.objects.create(sub_category=sub, major_category=name, code=code)
        else:
            Species.objects.create(sub_category=f'OTHER {name}', major_category=name, code=code or name[:3].upper())
        messages.success(request, f'Family "{name}" created with initial species.')
    return redirect('admin_manage_species')


@login_required
@admin_required
def admin_manage_family_species(request):
    """Admin: Manage species within a family."""
    family_name = request.GET.get('name', '')
    if not family_name:
        first = Species.objects.filter(major_category__iexact=request.GET.get('name', '')).first()
        return redirect('admin_manage_species')

    species_list = Species.objects.filter(major_category__iexact=family_name).order_by('sub_category')
    return render(request, 'admin/species/family_species.html', {
        'family_name': family_name,
        'species_list': species_list,
        'species_count': species_list.count(),
    })


@login_required
@admin_required
def admin_add_species(request, family_name):
    """Admin: Add a species to a family."""
    if request.method == 'POST':
        sub_category = request.POST.get('sub_category', '').strip()
        code = request.POST.get('code', '').strip()
        if not sub_category:
            messages.error(request, 'Species name is required.')
            return redirect('admin_manage_family_species_by_name', family_name=family_name)
        if Species.objects.filter(sub_category__iexact=sub_category, major_category__iexact=family_name).exists():
            messages.error(request, f'Species "{sub_category}" already exists in {family_name}.')
            return redirect('admin_manage_family_species_by_name', family_name=family_name)
        Species.objects.create(sub_category=sub_category, major_category=family_name, code=code)
        messages.success(request, f'Species "{sub_category}" added to {family_name}.')
    return redirect('admin_manage_family_species_by_name', family_name=family_name)


@login_required
@admin_required
def admin_edit_species(request, species_id):
    """Admin: Edit a species."""
    species = get_object_or_404(Species, id=species_id)
    family_name = species.major_category
    if request.method == 'POST':
        sub_category = request.POST.get('sub_category', '').strip()
        code = request.POST.get('code', '').strip()
        if not sub_category:
            messages.error(request, 'Species name is required.')
            return redirect('admin_manage_family_species_by_name', family_name=family_name)
        if Species.objects.filter(sub_category__iexact=sub_category, major_category__iexact=family_name).exclude(id=species_id).exists():
            messages.error(request, f'Species "{sub_category}" already exists in {family_name}.')
            return redirect('admin_manage_family_species_by_name', family_name=family_name)
        species.sub_category = sub_category
        species.code = code
        species.save()
        messages.success(request, f'Species updated to "{sub_category}".')
    return redirect('admin_manage_family_species_by_name', family_name=family_name)


@login_required
@admin_required
def admin_delete_species(request, species_id):
    """Admin: Delete a species."""
    species = get_object_or_404(Species, id=species_id)
    family_name = species.major_category
    if request.method == 'POST':
        name = str(species)
        has_approved = TransectSpecies.objects.filter(
            species=species, transect__assessment__status='approved'
        ).exists()
        if has_approved:
            messages.error(request, f'Cannot delete "{name}" — it is used in an approved assessment.')
            return redirect('admin_manage_family_species_by_name', family_name=family_name)
        species.delete()
        messages.success(request, f'Species "{name}" deleted.')
    return redirect('admin_manage_family_species_by_name', family_name=family_name)


@login_required
@admin_required
def admin_delete_family(request, family_name):
    """Admin: Delete an entire family and all its species."""
    if request.method == 'POST':
        species_in_family = Species.objects.filter(major_category__iexact=family_name)
        has_approved = TransectSpecies.objects.filter(
            species__in=species_in_family, transect__assessment__status='approved'
        ).exists()
        if has_approved:
            messages.error(request, f'Cannot delete family "{family_name}" — some species are used in approved assessments.')
            return redirect('admin_manage_species')
        count = species_in_family.count()
        species_in_family.delete()
        messages.success(request, f'Family "{family_name}" and {count} species deleted.')
    return redirect('admin_manage_species')


# ==================== CURATOR ASSESSMENT REVIEW VIEWS ====================

@login_required
@curator_required
def curator_assessments(request):
    """Curator: List all assessments with filters."""
    assessments = Assessment.objects.select_related(
        'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
    ).prefetch_related('transects').all()

    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    if status_filter:
        assessments = assessments.filter(status=status_filter)
    if search_query:
        assessments = assessments.filter(
            Q(barangay__name__icontains=search_query) |
            Q(municipality__name__icontains=search_query) |
            Q(uploaded_by__username__icontains=search_query)
        )

    stats = {
        'total': Assessment.objects.count(),
        'submitted': Assessment.objects.filter(status='submitted').count(),
        'approved': Assessment.objects.filter(status='approved').count(),
        'rejected': Assessment.objects.filter(status='rejected').count(),
        'draft': Assessment.objects.filter(status='draft').count(),
    }

    context = {
        'assessments': assessments,
        'status_filter': status_filter,
        'search_query': search_query,
        'stats': stats,
    }
    return render(request, 'curator/assessments.html', context)


@login_required
@curator_required
def curator_assessment_detail(request, assessment_id):
    """Curator: View assessment details with duplicate detection."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects__species_data__species', 'images'),
        id=assessment_id,
    )
    transects = assessment.transects.prefetch_related('species_data__species').all()

    # Parse Excel files on-the-fly if no species records yet
    parsed_species = []
    has_species_records = TransectSpecies.objects.filter(transect__assessment=assessment).exists()
    if not has_species_records:
        for transect in transects:
            if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
                sp, _ = parse_cpc_excel(transect.shallow_excel.path)
                for s in sp:
                    s['depth'] = 'Shallow'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)
            if transect.deep_excel and os.path.exists(transect.deep_excel.path):
                sp, _ = parse_cpc_excel(transect.deep_excel.path)
                for s in sp:
                    s['depth'] = 'Deep'
                    s['transect_id'] = transect.id
                    s['transect_number'] = transect.transect_number
                    parsed_species.append(s)

    # Find potential duplicate assessments
    similar_assessments = []
    current_species = set()

    if not has_species_records:
        for sp in parsed_species:
                current_species.add((sp['sub_category'], sp['major_category'], round(sp['cover'], 2)))
    else:
        for ts in TransectSpecies.objects.filter(
            transect__assessment=assessment
        ).select_related('species'):
            current_species.add((ts.species.sub_category, ts.species.major_category, float(ts.cover)))

    if current_species:
        other_assessments = Assessment.objects.filter(
            barangay=assessment.barangay,
            status='approved',
        ).exclude(id=assessment.id).order_by('-assessment_date')[:10]

        for other in other_assessments:
            other_species = set()
            for ts in TransectSpecies.objects.filter(
                transect__assessment=other
            ).select_related('species'):
                other_species.add((ts.species.sub_category, ts.species.major_category, float(ts.cover)))

            if not other_species:
                continue

            overlap = current_species & other_species
            total = max(len(current_species), len(other_species))
            if total == 0:
                continue

            match_pct = (len(overlap) / total) * 100
            if match_pct >= 50:
                similar_assessments.append({
                    'assessment': other,
                    'match_count': len(overlap),
                    'total': total,
                    'match_pct': round(match_pct, 1),
                    'is_exact': match_pct == 100 and len(current_species) == len(other_species),
                })

    return render(request, 'curator/assessment_detail.html', {
        'assessment': assessment,
        'transects': transects,
        'similar_assessments': similar_assessments,
        'parsed_species': parsed_species,
        'has_species_records': has_species_records,
    })


@login_required
@curator_required
def curator_confirm_approval(request, assessment_id):
    """Curator: Show approval warnings before confirming."""
    assessment = get_object_or_404(
        Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by'
        ),
        id=assessment_id,
        status='submitted',
    )
    transects = assessment.transects.all()

    all_species = []
    for transect in transects:
        if transect.shallow_excel and os.path.exists(transect.shallow_excel.path):
            species_list, _ = parse_cpc_excel(transect.shallow_excel.path)
            for sp in species_list:
                sp['transect'] = transect.transect_number
                sp['depth'] = 'Shallow'
                all_species.append(sp)
        if transect.deep_excel and os.path.exists(transect.deep_excel.path):
            species_list, _ = parse_cpc_excel(transect.deep_excel.path)
            for sp in species_list:
                sp['transect'] = transect.transect_number
                sp['depth'] = 'Deep'
                all_species.append(sp)

    current_set = {(sp['sub_category'], sp['major_category'], round(sp['cover'], 2)) for sp in all_species}
    duplicate_warnings = []
    if current_set:
        approved = Assessment.objects.filter(
            barangay=assessment.barangay, status='approved'
        ).exclude(id=assessment.id).order_by('-assessment_date')[:10]
        for other in approved:
            other_species = TransectSpecies.objects.filter(
                transect__assessment=other
            ).select_related('species')
            other_set = {(ts.species.sub_category, ts.species.major_category, round(float(ts.cover), 2)) for ts in other_species}
            if not other_set:
                continue
            overlap = current_set & other_set
            total = max(len(current_set), len(other_set))
            if total == 0:
                continue
            match_pct = (len(overlap) / total) * 100
            if match_pct >= 50:
                duplicate_warnings.append({
                    'assessment': other,
                    'match_count': len(overlap),
                    'total': total,
                    'match_pct': round(match_pct, 1),
                    'is_exact': match_pct == 100 and len(current_set) == len(other_set),
                    'overlap_species': list(overlap)[:5],
                })

    new_species = []
    for sp in all_species:
        exists = Species.objects.filter(
            sub_category__iexact=sp['sub_category'],
            major_category__iexact=sp['major_category'],
        ).exists()
        sp['is_new'] = not exists
        if not exists:
            new_species.append(sp)

    seen = set()
    unique_new = []
    for ns in new_species:
        key = (ns['sub_category'].lower(), ns['major_category'].lower())
        if key not in seen:
            seen.add(key)
            unique_new.append(ns)

    has_warnings = bool(duplicate_warnings or unique_new)

    return render(request, 'curator/confirm_approval.html', {
        'assessment': assessment,
        'transects': transects,
        'all_species': all_species,
        'duplicate_warnings': duplicate_warnings,
        'new_species': unique_new,
        'has_warnings': has_warnings,
    })


@login_required
@curator_required
def curator_assessment_action(request, assessment_id):
    """Curator: Approve, reject, or return assessment to pending."""
    with transaction.atomic():
        assessment = get_object_or_404(
            Assessment.objects.select_for_update(), id=assessment_id
        )

        action = request.POST.get('action')
        built_in = {val for val, label in Assessment.METHODOLOGY_CHOICES}

        if action == 'approve':
            assessment.status = 'approved'
            assessment.reviewed_by = request.user
            assessment.approved_at = timezone.now()
            assessment.save()
            species_count = populate_species_from_excel(assessment)
            save_barangay_transect_coords(assessment)
            if assessment.methodology not in built_in:
                CustomMethodology.objects.get_or_create(name=assessment.methodology)
            messages.success(request, f'Assessment #{assessment.id} approved. {species_count} species record(s) created.')
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '')
            assessment.status = 'rejected'
            assessment.reviewed_by = request.user
            assessment.approved_at = None
            assessment.notes = rejection_reason
            assessment.save()
            messages.info(request, f'Assessment #{assessment.id} has been rejected.')
        elif action == 'return_to_pending':
            assessment.status = 'submitted'
            assessment.reviewed_by = None
            assessment.approved_at = None
            assessment.notes = ''
            assessment.save()
            if assessment.methodology not in built_in:
                other_approved = Assessment.objects.filter(
                    status='approved', methodology=assessment.methodology
                ).exclude(id=assessment.id).exists()
                if not other_approved:
                    CustomMethodology.objects.filter(name=assessment.methodology).delete()
            cleanup_barangay_transect_coords(assessment)
            transects = assessment.transects.all()
            species_ids = list(TransectSpecies.objects.filter(
                transect__in=transects
            ).values_list('species_id', flat=True).distinct())
            deleted, _ = TransectSpecies.objects.filter(transect__in=transects).delete()
            if species_ids:
                other_approved_species = set(TransectSpecies.objects.filter(
                    transect__assessment__status='approved'
                ).exclude(transect__assessment=assessment).values_list('species_id', flat=True).distinct())
                orphan_ids = set(species_ids) - other_approved_species
                if orphan_ids:
                    Species.objects.filter(id__in=orphan_ids, source=Species.SOURCE_ASSESSMENT).delete()
            messages.success(request, f'Assessment #{assessment.id} returned to pending. {deleted} species record(s) removed.')
        else:
            messages.error(request, 'Invalid action.')

    return redirect('curator_assessments')


# ==================== SPECIES SUGGESTIONS API ====================

@login_required
def get_species_suggestions(request):
    """API: Get species that have been observed in approved assessments for a barangay.
    Only returns species from approved assessments."""
    barangay_id = request.GET.get('barangay_id')
    if not barangay_id:
        return JsonResponse({'species': []})

    species = Species.objects.filter(
        transect_data__transect__assessment__barangay_id=barangay_id,
        transect_data__transect__assessment__status='approved',
    ).distinct().values('id', 'sub_category', 'major_category', 'code')

    return JsonResponse({'species': list(species)})


# ==================== USER PROFILE VIEWS ====================

@login_required
def profile_view(request):
    """
    View for users to see their own profile
    """
    return render(request, 'profile/view.html', {'user': request.user})

@login_required
def profile_edit(request):
    """
    View for users to edit their own profile
    """
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('profile_view')
        else:
            error_list = []
            for field, errors in form.errors.items():
                label = form[field].label if field in form.fields else field.replace('_', ' ').title()
                for err in errors:
                    error_list.append(f"{label}: {err}")
            messages.error(request, '\n'.join(error_list) if error_list else 'Please correct the errors below.')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'profile/edit.html', {'form': form})

@login_required
def profile_change_password(request):
    """
    View for users to change their password
    """
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Update session to prevent logout
            update_session_auth_hash(request, user)
            messages.success(request, '✅ Your password has been changed successfully!')
            return redirect('profile_view')
        else:
            error_list = []
            for field, errors in form.errors.items():
                label = form[field].label if field in form.fields else field.replace('_', ' ').title()
                for err in errors:
                    error_list.append(f"{label}: {err}")
            messages.error(request, '\n'.join(error_list) if error_list else 'Please correct the errors below.')
    else:
        form = CustomPasswordChangeForm(request.user)
    
    return render(request, 'profile/change_password.html', {'form': form})


# ==================== PUBLIC MAP DASHBOARD ====================

def public_dashboard(request):
    """Public interactive map dashboard page (no login required)."""
    return render(request, 'public/dashboard.html', {'active_page': 'explore'})


def public_dashboard_data(request):
    """Public API: Return all approved assessment data for the map dashboard."""
    municipality_id = request.GET.get('municipality_id')
    barangay_id = request.GET.get('barangay_id')
    year_from = request.GET.get('year_from')
    year_to = request.GET.get('year_to')
    condition = request.GET.get('condition')

    assessments = Assessment.objects.filter(
        status='approved'
    ).select_related(
        'municipality', 'barangay', 'uploaded_by'
    ).prefetch_related(
        'transects__species_data__species'
    )

    if municipality_id:
        assessments = assessments.filter(municipality_id=municipality_id)
    if barangay_id:
        assessments = assessments.filter(barangay_id=barangay_id)
    if year_from:
        assessments = assessments.filter(assessment_date__year__gte=year_from)
    if year_to:
        assessments = assessments.filter(assessment_date__year__lte=year_to)
    if condition:
        assessments = assessments.filter(condition=condition)

    assessments = assessments.order_by('-assessment_date', '-approved_at')

    municipalities = {}
    total_transects = 0
    species_set = set()
    excellent = good = fair = poor = 0
    cover_sum = 0.0
    cover_count = 0
    trend_years = {}

    for a in assessments:
        m_id = a.municipality_id
        m_name = a.municipality.name
        b_id = a.barangay_id
        b_name = a.barangay.name

        if m_id not in municipalities:
            municipalities[m_id] = {
                'id': m_id, 'name': m_name, 'barangays': {}
            }
        if b_id not in municipalities[m_id]['barangays']:
            municipalities[m_id]['barangays'][b_id] = {
                'id': b_id, 'name': b_name, 'assessments': []
            }

        uploader_name = ''
        if a.uploaded_by:
            profile = getattr(a.uploaded_by, 'profile', None)
            if profile:
                uploader_name = profile.get_full_name()
            if not uploader_name:
                uploader_name = a.uploaded_by.get_full_name() or a.uploaded_by.email

        transects_list = []
        for t in a.transects.all():
            total_transects += 1
            sp = []
            for ts in t.species_data.all():
                species_set.add(ts.species.id)
                sp.append({
                    'co': ts.species.code,
                    'nm': ts.species.sub_category,
                    'fm': ts.species.major_category,
                    'cv': float(ts.cover),
                    'dp': ts.depth,
                })
            transects_list.append({
                'n': t.transect_number,
                's': {
                    's': [float(t.shallow_start_lat), float(t.shallow_start_lng)] if t.shallow_start_lat and t.shallow_start_lng else None,
                    'e': [float(t.shallow_end_lat), float(t.shallow_end_lng)] if t.shallow_end_lat and t.shallow_end_lng else None,
                },
                'd': {
                    's': [float(t.deep_start_lat), float(t.deep_start_lng)] if t.deep_start_lat and t.deep_start_lng else None,
                    'e': [float(t.deep_end_lat), float(t.deep_end_lng)] if t.deep_end_lat and t.deep_end_lng else None,
                },
                'sc': len(sp),
                'sp': sp,
            })

        cover = float(a.overall_coral_cover) if a.overall_coral_cover is not None else None

        municipalities[m_id]['barangays'][b_id]['assessments'].append({
            'id': a.id,
            'd': str(a.assessment_date),
            'c': a.condition or '',
            'cc': cover,
            'mt': a.get_methodology_display_name(),
            'up': uploader_name,
            'tr': transects_list,
        })

        if a.condition == 'excellent':
            excellent += 1
        elif a.condition == 'good':
            good += 1
        elif a.condition == 'fair':
            fair += 1
        elif a.condition == 'poor':
            poor += 1

        if cover is not None:
            cover_sum += cover
            cover_count += 1

        year = a.assessment_date.year
        if year not in trend_years:
            trend_years[year] = {'sum': 0.0, 'count': 0}
        if cover is not None:
            trend_years[year]['sum'] += cover
            trend_years[year]['count'] += 1

    result_m = []
    for m_id, m_data in municipalities.items():
        b_list = []
        for b_id, b_data in m_data['barangays'].items():
            b_as = b_data['assessments']
            lcv = None
            lc = ''
            for ba in b_as:
                if ba['cc'] is not None:
                    lcv = ba['cc']
                    lc = ba['c']
                    break

            all_lngs = []
            all_lats = []
            for ba in b_as:
                for bt in ba['tr']:
                    for depth_data in [bt['s'], bt['d']]:
                        if depth_data['s']:
                            all_lats.append(depth_data['s'][0])
                            all_lngs.append(depth_data['s'][1])
                        if depth_data['e']:
                            all_lats.append(depth_data['e'][0])
                            all_lngs.append(depth_data['e'][1])

            b_lat = round(sum(all_lats) / len(all_lats), 6) if all_lats else None
            b_lng = round(sum(all_lngs) / len(all_lngs), 6) if all_lngs else None

            b_list.append({
                'id': b_id,
                'name': b_data['name'],
                'lat': b_lat,
                'lng': b_lng,
                'lcv': lcv,
                'lc': lc,
                'ac': len(b_as),
                'a': b_as,
            })

        m_lngs = [b['lng'] for b in b_list if b['lng']]
        m_lats = [b['lat'] for b in b_list if b['lat']]
        m_lat = round(sum(m_lats) / len(m_lats), 6) if m_lats else None
        m_lng = round(sum(m_lngs) / len(m_lngs), 6) if m_lngs else None

        result_m.append({
            'id': m_id,
            'name': m_data['name'],
            'lat': m_lat,
            'lng': m_lng,
            'b': b_list,
        })

    result_t = []
    for year in sorted(trend_years.keys()):
        td = trend_years[year]
        result_t.append({
            'year': year,
            'avg_cover': round(td['sum'] / td['count'], 1) if td['count'] > 0 else None,
            'count': td['count'],
        })

    return JsonResponse({
        'm': result_m,
        's': {
            'total_assessments': len(assessments),
        'total_transects': total_transects,
        'avg_cover': round(cover_sum / cover_count, 1) if cover_count > 0 else 0,
        'total_species': len(species_set),
        'excellent': excellent,
        'good': good,
        'fair': fair,
        'poor': poor,
    },
    't': result_t,
})


@login_required
def assessments_sync(request):
    """Lightweight API: Return assessment IDs, statuses, and stats for polling.
    Returns a hash that changes when data changes, so clients only
    refetch the full table when something actually moved."""
    user_role = request.user.profile.role if hasattr(request.user, 'profile') else 'contributor'

    if user_role == 'contributor':
        assessments = Assessment.objects.filter(
            uploaded_by=request.user
        ).select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects')
    else:
        assessments = Assessment.objects.select_related(
            'municipality', 'barangay', 'uploaded_by', 'reviewed_by'
        ).prefetch_related('transects').all()

    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    if status_filter:
        assessments = assessments.filter(status=status_filter)
    if search_query:
        assessments = assessments.filter(
            Q(barangay__name__icontains=search_query) |
            Q(municipality__name__icontains=search_query) |
            Q(uploaded_by__username__icontains=search_query)
        )

    rows = []
    for a in assessments:
        rows.append({
            'id': a.id,
            'status': a.status,
            'status_display': a.get_status_display(),
            'barangay': a.barangay.name if a.barangay else '',
            'municipality': a.municipality.name if a.municipality else '',
            'date': a.assessment_date.strftime('%b %d, %Y') if a.assessment_date else '',
            'uploaded_by': a.uploaded_by.profile.get_full_name() if hasattr(a.uploaded_by, 'profile') and a.uploaded_by.profile.get_full_name() else (a.uploaded_by.email if a.uploaded_by else ''),
            'reviewed_by': (a.reviewed_by.profile.get_full_name() if hasattr(a.reviewed_by, 'profile') and a.reviewed_by.profile.get_full_name() else a.reviewed_by.email) if a.reviewed_by else None,
            'transect_count': a.transects.count(),
            'coral_cover': float(a.overall_coral_cover) if a.overall_coral_cover else None,
        })

    stats = {
        'total': Assessment.objects.count(),
        'submitted': Assessment.objects.filter(status='submitted').count(),
        'approved': Assessment.objects.filter(status='approved').count(),
        'rejected': Assessment.objects.filter(status='rejected').count(),
        'draft': Assessment.objects.filter(status='draft').count(),
    }

    import hashlib
    hash_input = json.dumps(rows, sort_keys=True, default=str)
    data_hash = hashlib.md5(hash_input.encode()).hexdigest()

    return JsonResponse({
        'hash': data_hash,
        'assessments': rows,
        'stats': stats,
        'role': user_role,
    })